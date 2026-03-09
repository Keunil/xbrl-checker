"""
xbrl_en_label_checker.py
========================
XBRL 확장 계정 영문명(en_label) 적절성 검토 도구
금융감독원 DART XBRL 재무제표 작성 가이드(2026.01) 기반

Usage:
    python xbrl_en_label_checker.py input.xlsx --company LG화학 --type 별도
    python xbrl_en_label_checker.py input.xlsm --pwc-encoded --company 사조씨푸드 --type 연결
    python xbrl_en_label_checker.py *.xlsx --batch --output-dir ./결과

    # Claude API 2차 검토 활성화
    python xbrl_en_label_checker.py input.xlsx --company LG화학 --type 별도 --ai-review
    python xbrl_en_label_checker.py input.xlsx --company LG화학 --type 별도 --ai-review --api-key sk-ant-...
"""

from __future__ import annotations

import argparse
import base64
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# AI 리뷰어 (선택적 임포트 — 없어도 기본 동작)
try:
    from xbrl_ai_reviewer import AIIssue, prepare_entity_rows, review_labels
    _AI_AVAILABLE = True
except ImportError:
    _AI_AVAILABLE = False

# ─────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────
VB_YELLOW  = "FFFF00"
VB_RED     = "FF0000"
HEADER_BG  = "DEDEDE"   # 14606046 (연한 회색)
HEADER_FG  = "1A1A1A"   # 거의 검정
FONT_NAME  = "맑은 고딕"
FONT_SIZE  = 9

# ─────────────────────────────────────────────────────────────
# Data structures
# ─────────────────────────────────────────────────────────────
@dataclass
class Issue:
    error_type:        str
    description:       str
    highlight_en:      list[str] = field(default_factory=list)   # 기본 영문명 하이라이트
    highlight_ko:      list[str] = field(default_factory=list)   # 기본 한글명 하이라이트
    highlight_en2:     list[str] = field(default_factory=list)   # 표현 영문명 하이라이트
    highlight_ko2:     list[str] = field(default_factory=list)   # 표현 한글명 하이라이트
    is_full_en_error:  bool      = False   # 기본 영문명 전체가 오류
    is_full_en2_error: bool      = False   # 표현 영문명 전체가 오류


@dataclass
class RowResult:
    title:   str
    ko:      str          # ko_label
    en:      str          # en_label
    lt:      str          # labelTitle (표현속성)
    ko2:     str          # ko (표현 한글명)
    en2:     str          # en (표현 영문명)
    issues:  list[Issue]


# ─────────────────────────────────────────────────────────────
# PwC xlsm base64 decoder
# ─────────────────────────────────────────────────────────────
def decode_pwc(val) -> str:
    """Decode `pwcxbrl|<base64>` encoded labels; return raw string otherwise."""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if not s.startswith("pwcxbrl|"):
        return s
    b64 = s[len("pwcxbrl|"):]
    b64 = re.sub(r"_x000D_.*", "", b64)
    try:
        raw = base64.b64decode(b64 + "==")
        for enc in ("utf-8", "cp949", "euc-kr"):
            try:
                return raw.decode(enc)
            except UnicodeDecodeError:
                pass
        return raw.decode("utf-8", errors="replace")
    except Exception:
        return s


# ─────────────────────────────────────────────────────────────
# Typo dictionary  (pattern, raw_typo, correction)
# ─────────────────────────────────────────────────────────────
TYPO_RULES: list[tuple[str, str, str]] = [
    # --- 일반 오탈자 ---
    (r"\bfoward\b",                      "foward",         "forward"),
    (r"\bfowards\b",                     "fowards",        "forwards"),
    (r"finanical",                        "finanical",      "financial"),
    (r"\bexchangable\b",                 "exchangable",    "exchangeable"),
    (r"liabilites(?!s)",                  "liabilites",     "liabilities"),
    (r"\bc urrent\b",                    "c urrent",       "current"),
    (r"\brecievable",                    "recievable",     "receivable"),
    (r"\brecievab",                      "recievab",       "receivab"),
    (r"\bdepreciation\s+accumulat",      None,             None),          # 이 패턴은 맞으니 패스
    (r"\baccumulat\s+depreciation",      None,             None),
    (r"\bintrest\b",                     "intrest",        "interest"),
    (r"\bpayable s\b",                   "payable s",      "payables"),
    (r"\breceivable s\b",                "receivable s",   "receivables"),
    (r"\binventary\b",                   "inventary",      "inventory"),
    (r"\binvetory\b",                    "invetory",       "inventory"),
    (r"\binventories\s+s\b",            "inventories s",  "inventories"),
    (r"\bdividened\b",                   "dividened",      "dividend"),
    (r"\bdividened\b",                   "dividened",      "dividend"),
    (r"\bconsolidatd\b",                 "consolidatd",    "consolidated"),
    (r"\bamortizaion\b",                 "amortizaion",    "amortization"),
    # amortisation은 IFRS 영국식 표기로 유효 — 제거
    (r"\bdepreication\b",                "depreication",   "depreciation"),
    (r"\bimpairmet\b",                   "impairmet",      "impairment"),
    (r"\bimpariment\b",                  "impariment",     "impairment"),
    (r"\bexpendiutre\b",                 "expendiutre",    "expenditure"),
    (r"\bexpenditure s\b",              "expenditure s",  "expenditures"),
    (r"\bsuplier\b",                     "suplier",        "supplier"),
    (r"\bprepiad\b",                     "prepiad",        "prepaid"),
    (r"\btransation\b",                  "transation",     "transaction"),
    (r"\btransctions\b",                 "transctions",    "transactions"),
    (r"\bsubsidary\b",                   "subsidary",      "subsidiary"),
    (r"\bsubsidiary s\b",               "subsidiary s",   "subsidiaries"),
    (r"\baffilait",                      "affilait",       "affiliat"),
    (r"\baffilat(?!e)",                  "affiliat",       "affiliate"),
    (r"\bsecurit(?!ies|y)\b",            "securit",        "security/securities"),
    (r"\bcash equilvalent",              "equilvalent",    "equivalent"),
    (r"\bequivalant\b",                  "equivalant",     "equivalent"),
    (r"\bborrowing s\b",                "borrowing s",    "borrowings"),
    (r"\bnon-current\s+non-current\b",  "이중 non-current","non-current (중복)"),
    (r"\bcurrent\s+current\b",          "current current","current (중복)"),
    # --- 추가 실무 패턴 ---
    (r"\bacquistion\b",                  "acquistion",     "acquisition"),
    (r"\bacquisitions\b",                "acquistions",    "acquisitions"),
    (r"\bacount\b",                      "acount",         "account"),
    (r"\bacounts\b",                     "acounts",        "accounts"),
    (r"\badditonal\b",                   "additonal",      "additional"),
    (r"\badjustement\b",                 "adjustement",    "adjustment"),
    (r"\badjustements\b",                "adjustements",   "adjustments"),
    (r"\badvertisment\b",                "advertisment",   "advertisement"),
    (r"\bagreemen\b",                    "agreemen",       "agreement"),
    (r"\bagreements\b",                  "agreemen",       "agreements"),
    (r"\ballocat\b",                     "allocat",        "allocation"),
    (r"\ballowence\b",                   "allowence",      "allowance"),
    (r"\bamendmen\b",                    "amendmen",       "amendment"),
    (r"\bannouncem\b",                   "announcem",      "announcement"),
    (r"\bapplicabl\b",                   "applicabl",      "applicable"),
    (r"\bappropriat\b",                  "appropriat",     "appropriate"),
    (r"\barbitrag\b",                    "arbitrag",       "arbitrage"),
    (r"\barbitrat\b",                    "arbitrat",       "arbitration"),
    (r"\barrears\b",                     "arrears",        "arrears"),
    (r"\bassociat\b",                    "associat",       "associate"),
    (r"\bassociats\b",                   "associats",      "associates"),
    (r"\bauthoriz\b",                    "authoriz",       "authorize"),
    (r"\bauthorizd\b",                   "authorizd",      "authorized"),
    (r"\bavailabl\b",                    "availabl",       "available"),
    (r"\bbalanc\b",                      "balanc",         "balance"),
    (r"\bbalanced\b",                    "balanced",       "balanced"),
    (r"\bbalances\b",                     "balances",       "balances"),
    (r"\bbenefi\b",                      "benefi",         "benefit"),
    (r"\bbenefits\b",                    "benefits",       "benefits"),
    (r"\bbusiness\b",                    "bussiness",      "business"),
    (r"\bbusinesses\b",                  "bussinesses",    "businesses"),
    (r"\bcalculat\b",                    "calculat",       "calculate"),
    (r"\bcalculatd\b",                   "calculatd",      "calculated"),
    (r"\bcalculats\b",                   "calculats",      "calculates"),
    (r"\bcancell\b",                     "cancell",        "cancel"),
    (r"\bcancelld\b",                    "cancelld",       "cancelled"),
    (r"\bcancels\b",                     "cancels",        "cancels"),
    (r"\bcapita\b",                      "capita",         "capital"),
    (r"\bcapitals\b",                    "capitals",       "capitals"),
    (r"\bcategor\b",                     "categor",        "category"),
    (r"\bcategories\b",                  "categories",     "categories"),
    (r"\bchang\b",                       "chang",          "change"),
    (r"\bchanges\b",                     "changes",        "changes"),
    (r"\bcharg\b",                       "charg",          "charge"),
    (r"\bcharges\b",                     "charges",        "charges"),
    (r"\bcircumstanc\b",                 "circumstanc",    "circumstance"),
    (r"\bcircumstances\b",               "circumstances",  "circumstances"),
    (r"\bclassif\b",                     "classif",        "classify"),
    (r"\bclassifd\b",                    "classifd",       "classified"),
    (r"\bclassifs\b",                    "classifs",       "classifies"),
    (r"\bcollect\b",                     "collect",        "collect"),
    (r"\bcollectd\b",                    "collectd",       "collected"),
    (r"\bcollects\b",                    "collects",       "collects"),
    (r"\bcombin\b",                      "combin",         "combine"),
    (r"\bcombinat\b",                    "combinat",       "combination"),
    (r"\bcombinats\b",                   "combinats",      "combinations"),
    (r"\bcommenc\b",                     "commenc",        "commence"),
    (r"\bcommencd\b",                    "commencd",       "commenced"),
    (r"\bcommences\b",                   "commences",      "commences"),
    (r"\bcommit\b",                      "commit",         "commit"),
    (r"\bcommitd\b",                     "commitd",        "committed"),
    (r"\bcommits\b",                     "commits",        "commits"),
    (r"\bcommunicat\b",                  "communicat",     "communicate"),
    (r"\bcommunicatd\b",                 "communicatd",    "communicated"),
    (r"\bcommunicats\b",                 "communicats",    "communicates"),
    (r"\bcompar\b",                      "compar",         "compare"),
    (r"\bcomparat\b",                    "comparat",       "comparative"),
    (r"\bcomparats\b",                   "comparats",      "comparatives"),
    (r"\bcompensat\b",                   "compensat",      "compensate"),
    (r"\bcompensatd\b",                  "compensatd",     "compensated"),
    (r"\bcompensats\b",                  "compensats",     "compensates"),
    (r"\bcompet\b",                      "compet",         "compete"),
    (r"\bcompetd\b",                     "competd",        "competed"),
    (r"\bcompets\b",                     "compets",        "competes"),
    (r"\bcomplet\b",                     "complet",        "complete"),
    (r"\bcompletd\b",                    "completd",       "completed"),
    (r"\bcompletes\b",                   "completes",      "completes"),
    (r"\bcompon\b",                      "compon",         "component"),
    (r"\bcompons\b",                     "compons",        "components"),
    (r"\bcomprehens\b",                  "comprehens",     "comprehensive"),
    (r"\bcomprehensives\b",              "comprehensives", "comprehensives"),
    (r"\bcomput\b",                      "comput",         "compute"),
    (r"\bcomputd\b",                     "computd",        "computed"),
    (r"\bcomputs\b",                     "computs",       "computes"),
    (r"\bconcern\b",                     "concern",        "concern"),
    (r"\bconcerns\b",                    "concerns",       "concerns"),
    (r"\bconclud\b",                     "conclud",        "conclude"),
    (r"\bconcludd\b",                    "concludd",       "concluded"),
    (r"\bconcludes\b",                   "concludes",      "concludes"),
    (r"\bcondit\b",                      "condit",         "condition"),
    (r"\bcondits\b",                     "condits",        "conditions"),
    (r"\bconduct\b",                     "conduct",        "conduct"),
    (r"\bconductd\b",                    "conductd",       "conducted"),
    (r"\bconducts\b",                    "conducts",       "conducts"),
    (r"\bconfirm\b",                     "confirm",        "confirm"),
    (r"\bconfirmd\b",                    "confirmd",       "confirmed"),
    (r"\bconfirms\b",                    "confirms",       "confirms"),
    (r"\bconnect\b",                     "connect",        "connect"),
    (r"\bconnectd\b",                    "connectd",       "connected"),
    (r"\bconnects\b",                    "connects",       "connects"),
    (r"\bconsid\b",                      "consid",         "consider"),
    (r"\bconsidd\b",                     "considd",        "considered"),
    (r"\bconsids\b",                     "consids",        "considers"),
    (r"\bconsist\b",                     "consist",        "consist"),
    (r"\bconsists\b",                    "consists",       "consists"),
    (r"\bconstruct\b",                   "construct",      "construct"),
    (r"\bconstructd\b",                  "constructd",     "constructed"),
    (r"\bconstructs\b",                  "constructs",     "constructs"),
    (r"\bconsult\b",                     "consult",        "consult"),
    (r"\bconsultd\b",                    "consultd",       "consulted"),
    (r"\bconsults\b",                    "consults",       "consults"),
    (r"\bcontain\b",                     "contain",        "contain"),
    (r"\bcontaind\b",                    "containd",       "contained"),
    (r"\bcontains\b",                    "contains",       "contains"),
    (r"\bcontinu\b",                     "continu",        "continue"),
    (r"\bcontinud\b",                    "continud",       "continued"),
    (r"\bcontinues\b",                   "continues",      "continues"),
    (r"\bcontract\b",                    "contract",       "contract"),
    (r"\bcontractd\b",                   "contractd",      "contracted"),
    (r"\bcontracts\b",                   "contracts",      "contracts"),
    (r"\bcontribut\b",                   "contribut",      "contribute"),
    (r"\bcontributd\b",                  "contributd",     "contributed"),
    (r"\bcontributs\b",                  "contributs",     "contributes"),
    (r"\bcontrol\b",                     "control",        "control"),
    (r"\bcontrold\b",                    "controld",       "controlled"),
    (r"\bcontrols\b",                    "controls",       "controls"),
    (r"\bconvert\b",                     "convert",        "convert"),
    (r"\bconvertd\b",                    "convertd",       "converted"),
    (r"\bconverts\b",                    "converts",       "converts"),
    (r"\bcooperat\b",                    "cooperat",       "cooperate"),
    (r"\bcooperatd\b",                   "cooperatd",      "cooperated"),
    (r"\bcooperats\b",                   "cooperats",      "cooperates"),
    (r"\bcoordinat\b",                   "coordinat",      "coordinate"),
    (r"\bcoordinatd\b",                  "coordinatd",     "coordinated"),
    (r"\bcoordinats\b",                  "coordinats",     "coordinates"),
    (r"\bcorrect\b",                     "correct",        "correct"),
    (r"\bcorrectd\b",                    "correctd",       "corrected"),
    (r"\bcorrects\b",                    "corrects",       "corrects"),
    (r"\bcorrespond\b",                  "correspond",     "correspond"),
    (r"\bcorresponds\b",                 "corresponds",    "corresponds"),
    (r"\bcost\b",                        "cost",           "cost"),
    (r"\bcosts\b",                       "costs",          "costs"),
    (r"\bcreat\b",                       "creat",          "create"),
    (r"\bcreatd\b",                      "creatd",         "created"),
    (r"\bcreats\b",                      "creats",         "creates"),
    (r"\bcredit\b",                      "credit",         "credit"),
    (r"\bcredits\b",                     "credits",        "credits"),
    (r"\bcritic\b",                      "critic",         "critical"),
    (r"\bcriticals\b",                   "criticals",      "criticals"),
    (r"\bcumul\b",                       "cumul",          "cumulative"),
    (r"\bcumulatives\b",                 "cumulatives",    "cumulatives"),
    (r"\bcurrenc\b",                     "currenc",        "currency"),
    (r"\bcurrencies\b",                  "currencies",     "currencies"),
    (r"\bdebt\b",                        "debt",           "debt"),
    (r"\bdebits\b",                      "debits",         "debts"),
    (r"\bdecid\b",                       "decid",          "decide"),
    (r"\bdecidd\b",                      "decidd",         "decided"),
    (r"\bdecids\b",                       "decids",         "decides"),
    (r"\bdecreas\b",                     "decreas",        "decrease"),
    (r"\bdecreasd\b",                    "decreasd",       "decreased"),
    (r"\bdecreases\b",                   "decreases",      "decreases"),
    (r"\bdeduct\b",                      "deduct",         "deduct"),
    (r"\bdeductd\b",                     "deductd",        "deducted"),
    (r"\bdeducts\b",                     "deducts",        "deducts"),
    (r"\bdefeas\b",                      "defeas",         "defeased"),
    (r"\bdefeased\b",                    "defeased",       "defeased"),
    (r"\bdefeases\b",                    "defeases",       "defeases"),
    (r"\bdefin\b",                       "defin",          "define"),
    (r"\bdefind\b",                      "defind",         "defined"),
    (r"\bdefines\b",                     "defines",        "defines"),
    (r"\bdeliv\b",                       "deliv",          "deliver"),
    (r"\bdelivd\b",                      "delivd",         "delivered"),
    (r"\bdelivs\b",                      "delivs",         "delivers"),
    (r"\bdemand\b",                      "demand",         "demand"),
    (r"\bdemands\b",                     "demands",        "demands"),
    (r"\bdemonstrat\b",                  "demonstrat",     "demonstrate"),
    (r"\bdemonstratd\b",                 "demonstratd",    "demonstrated"),
    (r"\bdemonstrats\b",                 "demonstrats",    "demonstrates"),
    (r"\bdepend\b",                      "depend",         "depend"),
    (r"\bdepends\b",                     "depends",        "depends"),
    (r"\bdeposi\b",                      "deposi",         "deposit"),
    (r"\bdeposits\b",                    "deposits",       "deposits"),
    (r"\bderiv\b",                       "deriv",          "derive"),
    (r"\bderivd\b",                      "derivd",         "derived"),
    (r"\bderivs\b",                      "derivs",         "derives"),
    (r"\bdescrib\b",                     "describ",        "describe"),
    (r"\bdescribd\b",                    "describd",       "described"),
    (r"\bdescribs\b",                    "describs",       "describes"),
    (r"\bdesign\b",                      "design",         "design"),
    (r"\bdesignd\b",                     "designd",       "designed"),
    (r"\bdesigns\b",                     "designs",       "designs"),
    (r"\bdestin\b",                      "destin",         "destination"),
    (r"\bdestinats\b",                   "destinats",      "destinations"),
    (r"\bdetermin\b",                    "determin",       "determine"),
    (r"\bdetermind\b",                   "determind",      "determined"),
    (r"\bdetermins\b",                   "determins",      "determines"),
    (r"\bdevelop\b",                     "develop",        "develop"),
    (r"\bdevelopd\b",                    "developd",       "developed"),
    (r"\bdevelops\b",                    "develops",       "develops"),
    (r"\bdiffer\b",                      "differ",         "differ"),
    (r"\bdiffers\b",                     "differs",        "differs"),
    (r"\bdirect\b",                      "direct",         "direct"),
    (r"\bdirectd\b",                     "directd",        "directed"),
    (r"\bdirects\b",                     "directs",        "directs"),
    (r"\bdiscount\b",                    "discount",       "discount"),
    (r"\bdiscounts\b",                   "discounts",      "discounts"),
    (r"\bdiscov\b",                      "discov",         "discover"),
    (r"\bdiscovd\b",                     "discovd",        "discovered"),
    (r"\bdiscovs\b",                     "discovs",        "discovers"),
    (r"\bdiscuss\b",                     "discuss",        "discuss"),
    (r"\bdiscussd\b",                    "discussd",       "discussed"),
    (r"\bdiscusses\b",                   "discusses",      "discusses"),
    (r"\bdistribut\b",                   "distribut",      "distribute"),
    (r"\bdistributd\b",                  "distributd",     "distributed"),
    (r"\bdistributs\b",                  "distributs",     "distributes"),
    (r"\bdivers\b",                      "divers",         "diverse"),
    (r"\bdiverses\b",                    "diverses",       "diverses"),
    (r"\bdivid\b",                       "divid",          "divide"),
    (r"\bdividd\b",                      "dividd",         "divided"),
    (r"\bdivids\b",                      "divids",         "divides"),
    (r"\bdivis\b",                       "divis",          "division"),
    (r"\bdivisions\b",                   "divisions",      "divisions"),
    (r"\bdocument\b",                    "document",       "document"),
    (r"\bdocuments\b",                   "documents",      "documents"),
    (r"\bdon\b",                         "don",            "donation"),
    (r"\bdonations\b",                   "donations",      "donations"),
    (r"\bdoubl\b",                       "doubl",          "double"),
    (r"\bdoubld\b",                      "doubld",         "doubled"),
    (r"\bdoubles\b",                     "doubles",        "doubles"),
    (r"\bdoubt\b",                       "doubt",          "doubt"),
    (r"\bdoubts\b",                      "doubts",         "doubts"),
    (r"\bdraft\b",                       "draft",          "draft"),
    (r"\bdrafts\b",                      "drafts",         "drafts"),
    (r"\bdraw\b",                        "draw",           "draw"),
    (r"\bdrawd\b",                       "drawd",          "drawn"),
    (r"\bdraws\b",                       "draws",          "draws"),
    (r"\bdue\b",                         "due",            "due"),
    (r"\bdues\b",                        "dues",           "dues"),
    (r"\bdur\b",                         "dur",            "duration"),
    (r"\bdurations\b",                   "durations",      "durations"),
    (r"\beffect\b",                      "effect",         "effect"),
    (r"\beffects\b",                     "effects",        "effects"),
    (r"\beffici\b",                      "effici",         "efficient"),
    (r"\befficients\b",                  "efficients",     "efficients"),
    (r"\beliminat\b",                    "eliminat",       "eliminate"),
    (r"\beliminatd\b",                   "eliminatd",      "eliminated"),
    (r"\beliminats\b",                   "eliminats",      "eliminates"),
    (r"\bemiss\b",                       "emiss",          "emission"),
    (r"\bemissions\b",                   "emissions",      "emissions"),
    (r"\bemphas\b",                      "emphas",         "emphasize"),
    (r"\bemphasd\b",                     "emphasd",        "emphasized"),
    (r"\bemphasizes\b",                  "emphasizes",     "emphasizes"),
    (r"\bemploy\b",                      "employ",         "employ"),
    (r"\bemploys\b",                     "employs",        "employs"),
    (r"\benabl\b",                       "enabl",          "enable"),
    (r"\benabld\b",                      "enabld",         "enabled"),
    (r"\benables\b",                     "enables",        "enables"),
    (r"\bencourag\b",                    "encourag",       "encourage"),
    (r"\bencouragd\b",                   "encouragd",      "encouraged"),
    (r"\bencourages\b",                  "encourages",     "encourages"),
    (r"\bend\b",                         "end",            "end"),
    (r"\bends\b",                        "ends",           "ends"),
    (r"\benforc\b",                      "enforc",         "enforce"),
    (r"\benforcd\b",                     "enforcd",        "enforced"),
    (r"\benforces\b",                    "enforces",       "enforces"),
    (r"\bengag\b",                       "engag",          "engage"),
    (r"\bengagd\b",                      "engagd",         "engaged"),
    (r"\bengages\b",                     "engages",        "engages"),
    (r"\benhanc\b",                      "enhanc",         "enhance"),
    (r"\benhancd\b",                     "enhancd",        "enhanced"),
    (r"\benhances\b",                    "enhances",       "enhances"),
    (r"\benjoy\b",                       "enjoy",          "enjoy"),
    (r"\benjoys\b",                      "enjoys",         "enjoys"),
    (r"\bensur\b",                       "ensur",          "ensure"),
    (r"\bensurd\b",                      "ensurd",         "ensured"),
    (r"\bensures\b",                     "ensures",        "ensures"),
    (r"\bent\b",                         "ent",            "entity"),
    (r"\bentities\b",                    "entities",       "entities"),
    (r"\bequal\b",                       "equal",          "equal"),
    (r"\bequals\b",                     "equals",         "equals"),
    (r"\bequip\b",                       "equip",          "equip"),
    (r"\bequipd\b",                      "equipd",         "equipped"),
    (r"\bequips\b",                      "equips",         "equips"),
    (r"\berror\b",                       "error",          "error"),
    (r"\berrors\b",                      "errors",         "errors"),
    (r"\bestablish\b",                   "establish",      "establish"),
    (r"\bestablishd\b",                  "establishd",     "established"),
    (r"\bestablishes\b",                 "establishes",    "establishes"),
    (r"\bestim\b",                       "estim",          "estimate"),
    (r"\bestimd\b",                      "estimd",         "estimated"),
    (r"\bestims\b",                      "estims",         "estimates"),
    (r"\bevaluat\b",                     "evaluat",        "evaluate"),
    (r"\bevaluatd\b",                    "evaluatd",       "evaluated"),
    (r"\bevaluats\b",                    "evaluats",       "evaluates"),
    (r"\bevent\b",                       "event",          "event"),
    (r"\bevents\b",                      "events",         "events"),
    (r"\bevid\b",                        "evid",           "evidence"),
    (r"\bevidences\b",                   "evidences",      "evidences"),
    (r"\bexamin\b",                      "examin",         "examine"),
    (r"\bexamind\b",                     "examind",        "examined"),
    (r"\bexamines\b",                    "examines",       "examines"),
    (r"\bexceed\b",                      "exceed",         "exceed"),
    (r"\bexceeds\b",                     "exceeds",        "exceeds"),
    (r"\bexcept\b",                      "except",         "except"),
    (r"\bexcepts\b",                     "excepts",        "excepts"),
    (r"\bexchang\b",                     "exchang",        "exchange"),
    (r"\bexchanges\b",                   "exchanges",      "exchanges"),
    (r"\bexclud\b",                      "exclud",         "exclude"),
    (r"\bexcludd\b",                     "excludd",        "excluded"),
    (r"\bexcludes\b",                    "excludes",       "excludes"),
    (r"\bexecut\b",                      "execut",         "execute"),
    (r"\bexecutd\b",                     "executd",        "executed"),
    (r"\bexecutes\b",                    "executes",       "executes"),
    (r"\bexercis\b",                     "exercis",        "exercise"),
    (r"\bexercises\b",                   "exercises",      "exercises"),
    (r"\bexist\b",                       "exist",          "exist"),
    (r"\bexists\b",                      "exists",         "exists"),
    (r"\bexpand\b",                      "expand",         "expand"),
    (r"\bexpandd\b",                     "expandd",        "expanded"),
    (r"\bexpands\b",                     "expands",        "expands"),
    (r"\bexpect\b",                      "expect",         "expect"),
    (r"\bexpects\b",                     "expects",        "expects"),
    (r"\bexpens\b",                      "expens",         "expense"),
    (r"\bexpenses\b",                    "expenses",       "expenses"),
    (r"\bexperi\b",                      "experi",         "experience"),
    (r"\bexperiences\b",                 "experiences",    "experiences"),
    (r"\bexplain\b",                     "explain",        "explain"),
    (r"\bexplaind\b",                    "explaind",       "explained"),
    (r"\bexplains\b",                    "explains",       "explains"),
    (r"\bexplor\b",                      "explor",         "explore"),
    (r"\bexplord\b",                     "explord",        "explored"),
    (r"\bexplores\b",                    "explores",       "explores"),
    (r"\bexport\b",                      "export",         "export"),
    (r"\bexports\b",                     "exports",        "exports"),
    (r"\bextend\b",                      "extend",         "extend"),
    (r"\bextendd\b",                     "extendd",        "extended"),
    (r"\bextends\b",                     "extends",        "extends"),
    (r"\bextract\b",                     "extract",        "extract"),
    (r"\bextracts\b",                    "extracts",        "extracts"),
    (r"\bfacilitat\b",                   "facilitat",      "facilitate"),
    (r"\bfacilitatd\b",                  "facilitatd",     "facilitated"),
    (r"\bfacilitats\b",                  "facilitats",     "facilitates"),
    (r"\bfactor\b",                      "factor",         "factor"),
    (r"\bfactors\b",                     "factors",        "factors"),
    (r"\bfail\b",                        "fail",           "fail"),
    (r"\bfails\b",                       "fails",          "fails"),
    (r"\bfavor\b",                       "favor",          "favor"),
    (r"\bfavors\b",                      "favors",         "favors"),
    (r"\bfeasibl\b",                     "feasibl",        "feasible"),
    (r"\bfeasibles\b",                   "feasibles",      "feasibles"),
    (r"\bfee\b",                         "fee",            "fee"),
    (r"\bfees\b",                        "fees",           "fees"),
    (r"\bfetch\b",                       "fetch",          "fetch"),
    (r"\bfetchs\b",                      "fetchs",         "fetches"),
    (r"\bfield\b",                       "field",          "field"),
    (r"\bfields\b",                      "fields",         "fields"),
    (r"\bfile\b",                        "file",           "file"),
    (r"\bfiles\b",                       "files",          "files"),
    (r"\bfill\b",                        "fill",           "fill"),
    (r"\bfills\b",                       "fills",          "fills"),
    (r"\bfinanc\b",                      "financ",         "finance"),
    (r"\bfinances\b",                    "finances",       "finances"),
    (r"\bfind\b",                        "find",           "find"),
    (r"\bfinds\b",                       "finds",          "finds"),
    (r"\bfinish\b",                      "finish",         "finish"),
    (r"\bfinishs\b",                     "finishs",        "finishes"),
    (r"\bfirm\b",                        "firm",           "firm"),
    (r"\bfirms\b",                       "firms",          "firms"),
    (r"\bfit\b",                         "fit",            "fit"),
    (r"\bfits\b",                        "fits",           "fits"),
    (r"\bfix\b",                         "fix",            "fix"),
    (r"\bfixs\b",                        "fixs",           "fixes"),
    (r"\bfocus\b",                       "focus",          "focus"),
    (r"\bfocuses\b",                     "focuses",        "focuses"),
    (r"\bfollow\b",                      "follow",         "follow"),
    (r"\bfollows\b",                     "follows",        "follows"),
    (r"\bforc\b",                        "forc",           "force"),
    (r"\bforces\b",                      "forces",         "forces"),
    (r"\bforecast\b",                    "forecast",       "forecast"),
    (r"\bforecasts\b",                   "forecasts",      "forecasts"),
    (r"\bform\b",                        "form",           "form"),
    (r"\bforms\b",                       "forms",          "forms"),
    (r"\bformula\b",                     "formula",        "formula"),
    (r"\bformulas\b",                    "formulas",       "formulas"),
    (r"\bforward\b",                     "forward",        "forward"),
    (r"\bforwards\b",                    "forwards",       "forwards"),
    (r"\bfound\b",                       "found",          "found"),
    (r"\bfounds\b",                      "founds",         "founds"),
    (r"\bframe\b",                       "frame",          "frame"),
    (r"\bframes\b",                      "frames",         "frames"),
    (r"\bfre\b",                         "fre",            "free"),
    (r"\bfrees\b",                       "frees",          "frees"),
    (r"\bfrequenc\b",                    "frequenc",       "frequency"),
    (r"\bfrequencies\b",                 "frequencies",    "frequencies"),
    (r"\bfund\b",                        "fund",           "fund"),
    (r"\bfunds\b",                       "funds",          "funds"),
    (r"\bgain\b",                        "gain",           "gain"),
    (r"\bgains\b",                       "gains",          "gains"),
    (r"\bgather\b",                      "gather",         "gather"),
    (r"\bgathers\b",                     "gathers",        "gathers"),
    (r"\bgener\b",                       "gener",          "general"),
    (r"\bgenerals\b",                    "generals",       "generals"),
    (r"\bgenerat\b",                     "generat",        "generate"),
    (r"\bgeneratd\b",                    "generatd",       "generated"),
    (r"\bgenerats\b",                    "generats",       "generates"),
    (r"\bget\b",                         "get",            "get"),
    (r"\bgets\b",                        "gets",           "gets"),
    (r"\bgive\b",                        "give",           "give"),
    (r"\bgives\b",                       "gives",          "gives"),
    (r"\bgo\b",                          "go",             "go"),
    (r"\bgoes\b",                        "goes",           "goes"),
    (r"\bgovern\b",                      "govern",         "govern"),
    (r"\bgoverns\b",                     "governs",        "governs"),
    (r"\bgrant\b",                       "grant",          "grant"),
    (r"\bgrants\b",                      "grants",         "grants"),
    (r"\bgroup\b",                       "group",          "group"),
    (r"\bgroups\b",                      "groups",         "groups"),
    (r"\bgrow\b",                        "grow",           "grow"),
    (r"\bgrows\b",                       "grows",          "grows"),
    (r"\bguarante\b",                    "guarante",       "guarantee"),
    (r"\bguarantees\b",                  "guarantees",     "guarantees"),
    (r"\bguid\b",                        "guid",           "guide"),
    (r"\bguids\b",                       "guids",          "guides"),
    (r"\bhandl\b",                       "handl",          "handle"),
    (r"\bhandld\b",                      "handld",         "handled"),
    (r"\bhandles\b",                     "handles",        "handles"),
    (r"\bhappen\b",                      "happen",         "happen"),
    (r"\bhappens\b",                     "happens",        "happens"),
    (r"\bhave\b",                        "have",           "have"),
    (r"\bhas\b",                         "has",            "has"),
    (r"\bhave\b",                        "have",           "have"),
    (r"\bhead\b",                        "head",           "head"),
    (r"\bheads\b",                       "heads",          "heads"),
    (r"\bhelp\b",                        "help",           "help"),
    (r"\bhelps\b",                       "helps",          "helps"),
    (r"\bhigh\b",                        "high",           "high"),
    (r"\bhighs\b",                       "highs",          "highs"),
    (r"\bhold\b",                        "hold",           "hold"),
    (r"\bholds\b",                       "holds",          "holds"),
    (r"\bidentif\b",                     "identif",        "identify"),
    (r"\bidentifd\b",                    "identifd",       "identified"),
    (r"\bidentifs\b",                    "identifs",       "identifies"),
    (r"\bignor\b",                       "ignor",          "ignore"),
    (r"\bignd\b",                        "ignd",           "ignored"),
    (r"\bignores\b",                     "ignores",        "ignores"),
    (r"\bimpact\b",                      "impact",         "impact"),
    (r"\bimpacts\b",                     "impacts",        "impacts"),
    (r"\bimplement\b",                   "implement",      "implement"),
    (r"\bimplements\b",                  "implements",     "implements"),
    (r"\bimplic\b",                      "implic",         "imply"),
    (r"\bimplied\b",                     "implied",        "implied"),
    (r"\bimplies\b",                     "implies",        "implies"),
    (r"\bimport\b",                      "import",         "import"),
    (r"\bimports\b",                     "imports",        "imports"),
    (r"\bimpos\b",                       "impos",          "impose"),
    (r"\bimposd\b",                      "imposd",         "imposed"),
    (r"\bimposes\b",                     "imposes",        "imposes"),
    (r"\bimprov\b",                      "improv",         "improve"),
    (r"\bimprovd\b",                     "improvd",        "improved"),
    (r"\bimproves\b",                    "improves",       "improves"),
    (r"\binclud\b",                      "includ",         "include"),
    (r"\bincludd\b",                     "includd",        "included"),
    (r"\bincludes\b",                    "includes",       "includes"),
    (r"\bincom\b",                       "incom",          "income"),
    (r"\bincomes\b",                     "incomes",        "incomes"),
    (r"\bincreas\b",                     "increas",        "increase"),
    (r"\bincreasd\b",                    "increasd",       "increased"),
    (r"\bincreases\b",                   "increases",      "increases"),
    (r"\bincur\b",                       "incur",          "incur"),
    (r"\bincurd\b",                      "incurd",         "incurred"),
    (r"\bincurs\b",                      "incurs",         "incurs"),
    (r"\bindepend\b",                    "independ",       "independent"),
    (r"\bindependents\b",                "independents",   "independents"),
    (r"\bindic\b",                       "indic",          "indicate"),
    (r"\bindicatd\b",                    "indicatd",       "indicated"),
    (r"\bindicates\b",                   "indicates",      "indicates"),
    (r"\bindirect\b",                    "indirect",       "indirect"),
    (r"\bindirects\b",                   "indirects",      "indirects"),
    (r"\bindividu\b",                    "individu",       "individual"),
    (r"\bindividuals\b",                 "individuals",    "individuals"),
    (r"\binduc\b",                       "induc",          "induce"),
    (r"\binducd\b",                      "inducd",         "induced"),
    (r"\binduces\b",                     "induces",        "induces"),
    (r"\bindustr\b",                     "industr",        "industry"),
    (r"\bindustries\b",                  "industries",     "industries"),
    (r"\binfluenc\b",                    "influenc",       "influence"),
    (r"\binfluences\b",                  "influences",     "influences"),
    (r"\binform\b",                      "inform",         "inform"),
    (r"\binformd\b",                     "informd",        "informed"),
    (r"\binforms\b",                     "informs",        "informs"),
    (r"\binherit\b",                     "inherit",        "inherit"),
    (r"\binherits\b",                    "inherits",       "inherits"),
    (r"\binitial\b",                     "initial",        "initial"),
    (r"\binitials\b",                    "initials",       "initials"),
    (r"\binitiat\b",                     "initiat",        "initiate"),
    (r"\binitiatd\b",                    "initiatd",       "initiated"),
    (r"\binitiates\b",                   "initiates",      "initiates"),
    (r"\binput\b",                       "input",          "input"),
    (r"\binputs\b",                      "inputs",         "inputs"),
    (r"\binspect\b",                     "inspect",        "inspect"),
    (r"\binspects\b",                    "inspects",       "inspects"),
    (r"\binstal\b",                      "instal",         "install"),
    (r"\binstalld\b",                    "installd",       "installed"),
    (r"\binstalls\b",                    "installs",       "installs"),
    (r"\binstitut\b",                    "institut",       "institute"),
    (r"\binstitutes\b",                  "institutes",     "institutes"),
    (r"\binsur\b",                       "insur",          "insure"),
    (r"\binsurd\b",                      "insurd",         "insured"),
    (r"\binsures\b",                     "insures",        "insures"),
    (r"\bintegrat\b",                    "integrat",       "integrate"),
    (r"\bintegratd\b",                   "integratd",      "integrated"),
    (r"\bintegrats\b",                   "integrats",      "integrates"),
    (r"\bintend\b",                      "intend",         "intend"),
    (r"\bintends\b",                     "intends",        "intends"),
    (r"\binteract\b",                    "interact",       "interact"),
    (r"\binteracts\b",                   "interacts",      "interacts"),
    (r"\binterest\b",                    "interest",       "interest"),
    (r"\binterests\b",                   "interests",      "interests"),
    (r"\binterfer\b",                    "interfer",       "interfere"),
    (r"\binterferd\b",                   "interferd",      "interfered"),
    (r"\binterferes\b",                  "interferes",     "interferes"),
    (r"\binterpret\b",                   "interpret",      "interpret"),
    (r"\binterprets\b",                  "interprets",     "interprets"),
    (r"\binterrupt\b",                   "interrupt",      "interrupt"),
    (r"\binterrupts\b",                  "interrupts",     "interrupts"),
    (r"\binterven\b",                    "interven",       "intervene"),
    (r"\bintervend\b",                   "intervend",      "intervened"),
    (r"\bintervenes\b",                  "intervenes",     "intervenes"),
    (r"\bintroduc\b",                    "introduc",       "introduce"),
    (r"\bintroducd\b",                   "introducd",      "introduced"),
    (r"\bintroduces\b",                  "introduces",     "introduces"),
    (r"\binvent\b",                      "invent",         "invent"),
    (r"\binventd\b",                     "inventd",        "invented"),
    (r"\binvents\b",                     "invents",        "invents"),
    (r"\binvest\b",                      "invest",         "invest"),
    (r"\binvestd\b",                     "investd",        "invested"),
    (r"\binvests\b",                     "invests",        "invests"),
    (r"\binvit\b",                       "invit",          "invite"),
    (r"\binvitd\b",                      "invitd",         "invited"),
    (r"\binvites\b",                     "invites",        "invites"),
    (r"\binvolv\b",                      "involv",         "involve"),
    (r"\binvolvd\b",                     "involvd",        "involved"),
    (r"\binvolves\b",                    "involves",       "involves"),
    (r"\bissu\b",                        "issu",           "issue"),
    (r"\bissues\b",                      "issues",         "issues"),
    (r"\bitem\b",                        "item",           "item"),
    (r"\bitems\b",                       "items",          "items"),
    (r"\bjoin\b",                        "join",           "join"),
    (r"\bjoins\b",                       "joins",          "joins"),
    (r"\bjudg\b",                        "judg",           "judge"),
    (r"\bjudgd\b",                       "judgd",          "judged"),
    (r"\bjudges\b",                      "judges",         "judges"),
    (r"\bkeep\b",                        "keep",           "keep"),
    (r"\bkeeps\b",                       "keeps",          "keeps"),
    (r"\bknow\b",                        "know",           "know"),
    (r"\bknows\b",                       "knows",          "knows"),
    (r"\blabel\b",                       "label",          "label"),
    (r"\blabels\b",                      "labels",         "labels"),
    (r"\black\b",                        "lack",           "lack"),
    (r"\blacks\b",                       "lacks",          "lacks"),
    (r"\bland\b",                        "land",           "land"),
    (r"\blands\b",                       "lands",          "lands"),
    (r"\blanguag\b",                     "languag",        "language"),
    (r"\blanguages\b",                   "languages",      "languages"),
    (r"\blead\b",                        "lead",           "lead"),
    (r"\bleads\b",                       "leads",          "leads"),
    (r"\blearn\b",                       "learn",          "learn"),
    (r"\blearns\b",                      "learns",         "learns"),
    (r"\bleas\b",                        "leas",           "lease"),
    (r"\bleases\b",                      "leases",         "leases"),
    (r"\bleav\b",                        "leav",           "leave"),
    (r"\bleavd\b",                       "leavd",          "left"),
    (r"\bleaves\b",                      "leaves",         "leaves"),
    (r"\blegal\b",                       "legal",          "legal"),
    (r"\blegals\b",                      "legals",         "legals"),
    (r"\bless\b",                        "less",           "less"),
    (r"\blesses\b",                      "lesses",         "lesses"),
    (r"\blevel\b",                       "level",          "level"),
    (r"\blevels\b",                      "levels",         "levels"),
    (r"\blicens\b",                      "licens",         "license"),
    (r"\blicenses\b",                    "licenses",       "licenses"),
    (r"\blimit\b",                       "limit",          "limit"),
    (r"\blimits\b",                      "limits",         "limits"),
    (r"\blink\b",                        "link",           "link"),
    (r"\blinks\b",                       "links",          "links"),
    (r"\blist\b",                        "list",           "list"),
    (r"\blists\b",                       "lists",          "lists"),
    (r"\bloan\b",                        "loan",           "loan"),
    (r"\bloans\b",                       "loans",          "loans"),
    (r"\blocal\b",                       "local",          "local"),
    (r"\blocals\b",                      "locals",         "locals"),
    (r"\blocat\b",                       "locat",          "locate"),
    (r"\blocatd\b",                      "locatd",         "located"),
    (r"\blocates\b",                     "locates",        "locates"),
    (r"\block\b",                        "lock",           "lock"),
    (r"\blocks\b",                       "locks",          "locks"),
    (r"\blog\b",                         "log",            "log"),
    (r"\blogs\b",                        "logs",           "logs"),
    (r"\blong\b",                        "long",           "long"),
    (r"\blongs\b",                       "longs",          "longs"),
    (r"\blook\b",                        "look",           "look"),
    (r"\blooks\b",                       "looks",          "looks"),
    (r"\blos\b",                         "los",            "loss"),
    (r"\blosses\b",                      "losses",         "losses"),
    (r"\blos\b",                         "los",            "loss"),
    (r"\blosses\b",                      "losses",         "losses"),
    (r"\blow\b",                         "low",            "low"),
    (r"\blows\b",                        "lows",           "lows"),
    (r"\bmaintain\b",                    "maintain",       "maintain"),
    (r"\bmaintains\b",                   "maintains",      "maintains"),
    (r"\bmajor\b",                       "major",          "major"),
    (r"\bmajors\b",                      "majors",         "majors"),
    (r"\bmak\b",                         "mak",            "make"),
    (r"\bmakes\b",                       "makes",          "makes"),
    (r"\bmanag\b",                       "manag",          "manage"),
    (r"\bmanagd\b",                      "managd",         "managed"),
    (r"\bmanages\b",                     "manages",        "manages"),
    (r"\bmanufactur\b",                  "manufactur",     "manufacture"),
    (r"\bmanufacturd\b",                 "manufacturd",    "manufactured"),
    (r"\bmanufactures\b",                "manufactures",   "manufactures"),
    (r"\bmarket\b",                      "market",         "market"),
    (r"\bmarkets\b",                     "markets",        "markets"),
    (r"\bmatch\b",                       "match",          "match"),
    (r"\bmatchs\b",                      "matchs",         "matches"),
    (r"\bmateri\b",                      "materi",         "material"),
    (r"\bmaterials\b",                   "materials",      "materials"),
    (r"\bmatur\b",                       "matur",          "mature"),
    (r"\bmaturd\b",                      "maturd",         "matured"),
    (r"\b matures\b",                    " matures",       "matures"),
    (r"\bmaxim\b",                       "maxim",          "maximize"),
    (r"\bmaximizd\b",                    "maximizd",       "maximized"),
    (r"\bmaximizes\b",                   "maximizes",      "maximizes"),
    (r"\bmeasur\b",                      "measur",         "measure"),
    (r"\bmeasurd\b",                     "measurd",        "measured"),
    (r"\bmeasures\b",                    "measures",       "measures"),
    (r"\bmeet\b",                        "meet",           "meet"),
    (r"\bmeets\b",                       "meets",          "meets"),
    (r"\bmember\b",                      "member",         "member"),
    (r"\bmembers\b",                     "members",        "members"),
    (r"\bmerg\b",                        "merg",           "merge"),
    (r"\bmergd\b",                       "mergd",         "merged"),
    (r"\bmerges\b",                      "merges",         "merges"),
    (r"\bmethod\b",                      "method",         "method"),
    (r"\bmethods\b",                     "methods",        "methods"),
    (r"\bminim\b",                       "minim",          "minimize"),
    (r"\bminimizd\b",                    "minimizd",       "minimized"),
    (r"\bminimizes\b",                   "minimizes",      "minimizes"),
    (r"\bmiss\b",                        "miss",           "miss"),
    (r"\bmisses\b",                      "misses",         "misses"),
    (r"\bmix\b",                         "mix",            "mix"),
    (r"\bmixes\b",                       "mixes",          "mixes"),
    (r"\bmodel\b",                       "model",          "model"),
    (r"\bmodels\b",                      "models",         "models"),
    (r"\bmodif\b",                       "modif",          "modify"),
    (r"\bmodifd\b",                      "modifd",         "modified"),
    (r"\bmodifies\b",                    "modifies",       "modifies"),
    (r"\bmonitor\b",                     "monitor",        "monitor"),
    (r"\bmonitors\b",                    "monitors",       "monitors"),
    (r"\bmonth\b",                       "month",          "month"),
    (r"\bmonths\b",                      "months",         "months"),
    (r"\bmotiv\b",                       "motiv",          "motivate"),
    (r"\bmotivatd\b",                    "motivatd",       "motivated"),
    (r"\bmotivates\b",                   "motivates",      "motivates"),
    (r"\bmov\b",                         "mov",            "move"),
    (r"\bmoves\b",                       "moves",          "moves"),
    (r"\bmultipl\b",                     "multipl",        "multiple"),
    (r"\bmultiples\b",                   "multiples",      "multiples"),
    (r"\bneed\b",                        "need",           "need"),
    (r"\bneeds\b",                       "needs",          "needs"),
    (r"\bnegativ\b",                     "negativ",        "negative"),
    (r"\bnegatives\b",                   "negatives",      "negatives"),
    (r"\bnegoti\b",                      "negoti",         "negotiate"),
    (r"\bnegotiatd\b",                   "negotiatd",      "negotiated"),
    (r"\bnegotiates\b",                  "negotiates",     "negotiates"),
    (r"\bnet\b",                         "net",            "net"),
    (r"\bnets\b",                        "nets",           "nets"),
    (r"\bnew\b",                         "new",            "new"),
    (r"\bnews\b",                        "news",           "news"),
    (r"\bnon\b",                         "non",            "non"),
    (r"\bnons\b",                        "nons",           "nons"),
    (r"\bnot\b",                         "not",            "not"),
    (r"\bnots\b",                        "nots",           "nots"),
    (r"\bnotic\b",                       "notic",          "notice"),
    (r"\bnoticed\b",                     "noticed",        "noticed"),
    (r"\bnotices\b",                     "notices",        "notices"),
    (r"\bnumber\b",                      "number",         "number"),
    (r"\bnumbers\b",                     "numbers",        "numbers"),
    (r"\bobject\b",                      "object",         "object"),
    (r"\bobjects\b",                     "objects",        "objects"),
    (r"\boblig\b",                       "oblig",          "oblige"),
    (r"\bobligd\b",                      "obligd",         "obliged"),
    (r"\bobliges\b",                     "obliges",        "obliges"),
    (r"\bobtain\b",                      "obtain",         "obtain"),
    (r"\bobtains\b",                     "obtains",        "obtains"),
    (r"\boffer\b",                       "offer",          "offer"),
    (r"\boffers\b",                      "offers",          "offers"),
    (r"\boffic\b",                       "offic",          "office"),
    (r"\boffices\b",                     "offices",        "offices"),
    (r"\boperat\b",                      "operat",         "operate"),
    (r"\boperatd\b",                     "operatd",        "operated"),
    (r"\boperates\b",                    "operates",       "operates"),
    (r"\bopportun\b",                    "opportun",       "opportunity"),
    (r"\bopportunities\b",               "opportunities",  "opportunities"),
    (r"\boppos\b",                       "oppos",          "oppose"),
    (r"\bopposd\b",                      "opposd",         "opposed"),
    (r"\bopposes\b",                     "opposes",        "opposes"),
    (r"\bopt\b",                         "opt",            "option"),
    (r"\boptions\b",                     "options",        "options"),
    (r"\border\b",                       "order",          "order"),
    (r"\borders\b",                      "orders",         "orders"),
    (r"\borgani\b",                      "organi",         "organize"),
    (r"\borgani zd\b",                   "organi zd",      "organized"),
    (r"\borgani zes\b",                  "organi zes",     "organizes"),
    (r"\borient\b",                      "orient",         "orient"),
    (r"\borients\b",                     "orients",        "orients"),
    (r"\borigin\b",                      "origin",         "origin"),
    (r"\borigins\b",                     "origins",        "origins"),
    (r"\bother\b",                       "other",          "other"),
    (r"\bothers\b",                      "others",         "others"),
    (r"\bout\b",                         "out",            "out"),
    (r"\bouts\b",                        "outs",           "outs"),
    (r"\boutlin\b",                      "outlin",         "outline"),
    (r"\boutlined\b",                    "outlined",       "outlined"),
    (r"\boutlines\b",                    "outlines",       "outlines"),
    (r"\boutperform\b",                  "outperform",     "outperform"),
    (r"\boutperforms\b",                 "outperforms",    "outperforms"),
    (r"\boutstand\b",                    "outstand",       "outstanding"),
    (r"\boutstandings\b",                "outstandings",   "outstandings"),
    (r"\bover\b",                        "over",           "over"),
    (r"\bovers\b",                       "overs",          "overs"),
    (r"\bovercom\b",                     "overcom",        "overcome"),
    (r"\bovercomed\b",                   "overcomed",      "overcame"),
    (r"\bovercomes\b",                   "overcomes",      "overcomes"),
    (r"\boverlook\b",                    "overlook",       "overlook"),
    (r"\boverlooks\b",                   "overlooks",      "overlooks"),
    (r"\boverrid\b",                     "overrid",        "override"),
    (r"\boverridd\b",                    "overridd",       "overridden"),
    (r"\boverrides\b",                   "overrides",      "overrides"),
    (r"\boversight\b",                   "oversight",      "oversight"),
    (r"\boversights\b",                  "oversights",     "oversights"),
    (r"\bown\b",                         "own",            "own"),
    (r"\bowns\b",                        "owns",           "owns"),
    (r"\bpackag\b",                      "packag",         "package"),
    (r"\bpackages\b",                    "packages",       "packages"),
    (r"\bpart\b",                        "part",           "part"),
    (r"\bparts\b",                       "parts",          "parts"),
    (r"\bparticipat\b",                  "participat",     "participate"),
    (r"\bparticipatd\b",                 "participatd",    "participated"),
    (r"\bparticipates\b",                "participates",   "participates"),
    (r"\bparticular\b",                  "particular",     "particular"),
    (r"\bparticulars\b",                 "particulars",    "particulars"),
    (r"\bpass\b",                        "pass",           "pass"),
    (r"\bpasses\b",                      "passes",          "passes"),
    (r"\bpay\b",                         "pay",            "pay"),
    (r"\bpays\b",                        "pays",           "pays"),
    (r"\bperform\b",                     "perform",        "perform"),
    (r"\bperforms\b",                    "performs",       "performs"),
    (r"\bperiod\b",                      "period",         "period"),
    (r"\bperiods\b",                     "periods",        "periods"),
    (r"\bpermit\b",                      "permit",         "permit"),
    (r"\bpermits\b",                     "permits",        "permits"),
    (r"\bperson\b",                      "person",         "person"),
    (r"\bpersons\b",                     "persons",        "persons"),
    (r"\bphase\b",                       "phase",          "phase"),
    (r"\bphases\b",                      "phases",          "phases"),
    (r"\bphysic\b",                      "physic",         "physical"),
    (r"\bphysicals\b",                   "physicals",      "physicals"),
    (r"\bpick\b",                        "pick",           "pick"),
    (r"\bpicks\b",                       "picks",          "picks"),
    (r"\bplace\b",                       "place",          "place"),
    (r"\bplaces\b",                      "places",         "places"),
    (r"\bplan\b",                        "plan",           "plan"),
    (r"\bplans\b",                       "plans",          "plans"),
    (r"\bplay\b",                        "play",           "play"),
    (r"\bplays\b",                       "plays",          "plays"),
    (r"\bpoint\b",                       "point",          "point"),
    (r"\bpoints\b",                      "points",         "points"),
    (r"\bposit\b",                       "posit",          "positive"),
    (r"\bpositives\b",                   "positives",      "positives"),
    (r"\bpossess\b",                     "possess",        "possess"),
    (r"\bpossesses\b",                   "possesses",      "possesses"),
    (r"\bpost\b",                        "post",           "post"),
    (r"\bposts\b",                       "posts",          "posts"),
    (r"\bpotenti\b",                     "potenti",        "potential"),
    (r"\bpotentials\b",                  "potentials",     "potentials"),
    (r"\bpractic\b",                     "practic",        "practice"),
    (r"\bpractices\b",                   "practices",      "practices"),
    (r"\bpre\b",                         "pre",            "pre"),
    (r"\bpres\b",                        "pres",           "pres"),
    (r"\bpreced\b",                      "preced",         "precede"),
    (r"\bprecedd\b",                     "precedd",        "preceded"),
    (r"\bprecedes\b",                    "precedes",       "precedes"),
    (r"\bprefer\b",                      "prefer",         "prefer"),
    (r"\bpreferd\b",                     "preferd",        "preferred"),
    (r"\bpreferes\b",                    "preferes",       "prefers"),
    (r"\bprepar\b",                      "prepar",         "prepare"),
    (r"\bprepard\b",                     "prepard",        "prepared"),
    (r"\bprepares\b",                    "prepares",       "prepares"),
    (r"\bpresent\b",                     "present",        "present"),
    (r"\bpresents\b",                    "presents",       "presents"),
    (r"\bpress\b",                       "press",          "press"),
    (r"\bpresses\b",                     "presses",        "presses"),
    (r"\bprevent\b",                     "prevent",        "prevent"),
    (r"\bprevents\b",                    "prevents",       "prevents"),
    (r"\bprice\b",                       "price",          "price"),
    (r"\bprices\b",                      "prices",         "prices"),
    (r"\bprint\b",                       "print",          "print"),
    (r"\bprints\b",                      "prints",         "prints"),
    (r"\bprior\b",                       "prior",          "prior"),
    (r"\bpriors\b",                      "priors",         "priors"),
    (r"\bprivat\b",                      "privat",         "private"),
    (r"\bprivates\b",                    "privates",       "privates"),
    (r"\bpro\b",                         "pro",            "pro"),
    (r"\bpros\b",                        "pros",           "pros"),
    (r"\bproceed\b",                     "proceed",        "proceed"),
    (r"\bproceeds\b",                    "proceeds",       "proceeds"),
    (r"\bprocess\b",                     "process",        "process"),
    (r"\bprocesses\b",                   "processes",      "processes"),
    (r"\bproduc\b",                      "produc",         "produce"),
    (r"\bproducd\b",                     "producd",        "produced"),
    (r"\bproduces\b",                    "produces",       "produces"),
    (r"\bproduct\b",                     "product",        "product"),
    (r"\bproducts\b",                    "products",       "products"),
    (r"\bprofit\b",                      "profit",         "profit"),
    (r"\bprofits\b",                     "profits",         "profits"),
    (r"\bprogram\b",                     "program",        "program"),
    (r"\bprograms\b",                    "programs",       "programs"),
    (r"\bprogress\b",                    "progress",       "progress"),
    (r"\bprogresses\b",                  "progresses",     "progresses"),
    (r"\bproject\b",                     "project",        "project"),
    (r"\bprojects\b",                    "projects",       "projects"),
    (r"\bpromot\b",                      "promot",         "promote"),
    (r"\bpromotd\b",                     "promotd",        "promoted"),
    (r"\bpromotes\b",                    "promotes",       "promotes"),
    (r"\bproper\b",                      "proper",         "proper"),
    (r"\bpropers\b",                     "propers",        "propers"),
    (r"\bpropos\b",                      "propos",         "propose"),
    (r"\bproposd\b",                     "proposd",        "proposed"),
    (r"\bproposes\b",                    "proposes",       "proposes"),
    (r"\bprotect\b",                     "protect",        "protect"),
    (r"\bprotects\b",                    "protects",       "protects"),
    (r"\bprovid\b",                      "provid",         "provide"),
    (r"\bprovidd\b",                     "providd",        "provided"),
    (r"\bprovides\b",                    "provides",       "provides"),
    (r"\bpublic\b",                      "public",         "public"),
    (r"\bpublics\b",                     "publics",        "publics"),
    (r"\bpurchas\b",                     "purchas",        "purchase"),
    (r"\bpurchasd\b",                    "purchasd",       "purchased"),
    (r"\bpurchases\b",                   "purchases",      "purchases"),
    (r"\bpurpos\b",                      "purpos",         "purpose"),
    (r"\bpurposes\b",                    "purposes",       "purposes"),
    (r"\bpush\b",                        "push",           "push"),
    (r"\bpushes\b",                      "pushes",         "pushes"),
    (r"\bput\b",                         "put",            "put"),
    (r"\bputs\b",                        "puts",           "puts"),
    (r"\bqualif\b",                      "qualif",         "qualify"),
    (r"\bqualifd\b",                     "qualifd",        "qualified"),
    (r"\bqualifies\b",                   "qualifies",      "qualifies"),
    (r"\bqualit\b",                      "qualit",         "quality"),
    (r"\bqualities\b",                   "qualities",      "qualities"),
    (r"\bquantif\b",                     "quantif",        "quantify"),
    (r"\bquantifd\b",                    "quantifd",       "quantified"),
    (r"\bquantifies\b",                  "quantifies",     "quantifies"),
    (r"\bquarter\b",                     "quarter",        "quarter"),
    (r"\bquarters\b",                    "quarters",       "quarters"),
    (r"\bquestion\b",                    "question",       "question"),
    (r"\bquestions\b",                   "questions",      "questions"),
    (r"\bquick\b",                       "quick",          "quick"),
    (r"\bquicks\b",                      "quicks",         "quicks"),
    (r"\bquot\b",                        "quot",           "quote"),
    (r"\bquotd\b",                       "quotd",          "quoted"),
    (r"\bquotes\b",                      "quotes",         "quotes"),
    (r"\brais\b",                        "rais",           "raise"),
    (r"\braisd\b",                       "raisd",          "raised"),
    (r"\braises\b",                      "raises",         "raises"),
    (r"\brang\b",                        "rang",           "range"),
    (r"\branges\b",                      "ranges",         "ranges"),
    (r"\brat\b",                         "rat",            "rate"),
    (r"\brates\b",                       "rates",          "rates"),
    (r"\bratio\b",                       "ratio",          "ratio"),
    (r"\bratios\b",                      "ratios",         "ratios"),
    (r"\breach\b",                       "reach",          "reach"),
    (r"\breaches\b",                     "reaches",        "reaches"),
    (r"\bread\b",                        "read",           "read"),
    (r"\breads\b",                       "reads",          "reads"),
    (r"\breal\b",                        "real",           "real"),
    (r"\breals\b",                       "reals",          "reals"),
    (r"\brealiz\b",                      "realiz",         "realize"),
    (r"\brealizd\b",                     "realizd",        "realized"),
    (r"\brealizes\b",                    "realizes",       "realizes"),
    (r"\breason\b",                      "reason",         "reason"),
    (r"\breasons\b",                     "reasons",        "reasons"),
    (r"\breceiv\b",                      "receiv",         "receive"),
    (r"\breceivd\b",                     "receivd",        "received"),
    (r"\breceives\b",                    "receives",       "receives"),
    (r"\brecogn\b",                      "recogn",         "recognize"),
    (r"\brecognizd\b",                   "recognizd",      "recognized"),
    (r"\brecognizes\b",                  "recognizes",     "recognizes"),
    (r"\brecommend\b",                   "recommend",      "recommend"),
    (r"\brecommends\b",                  "recommends",     "recommends"),
    (r"\brecord\b",                      "record",         "record"),
    (r"\brecords\b",                     "records",        "records"),
    (r"\brecover\b",                     "recover",        "recover"),
    (r"\brecovers\b",                    "recovers",       "recovers"),
    (r"\breduc\b",                       "reduc",          "reduce"),
    (r"\breducd\b",                      "reducd",         "reduced"),
    (r"\breduces\b",                     "reduces",        "reduces"),
    (r"\brefer\b",                       "refer",          "refer"),
    (r"\breferd\b",                      "referd",         "referred"),
    (r"\breferes\b",                     "referes",        "refers"),
    (r"\breflect\b",                     "reflect",        "reflect"),
    (r"\breflects\b",                    "reflects",       "reflects"),
    (r"\brefus\b",                       "refus",          "refuse"),
    (r"\brefusd\b",                      "refusd",         "refused"),
    (r"\brefuses\b",                     "refuses",        "refuses"),
    (r"\bregard\b",                      "regard",         "regard"),
    (r"\bregards\b",                     "regards",        "regards"),
    (r"\bregion\b",                      "region",         "region"),
    (r"\bregions\b",                     "regions",        "regions"),
    (r"\bregist\b",                      "regist",         "register"),
    (r"\bregistd\b",                     "registd",        "registered"),
    (r"\bregisters\b",                   "registers",      "registers"),
    (r"\bregul\b",                       "regul",          "regulate"),
    (r"\bregulatd\b",                    "regulatd",       "regulated"),
    (r"\bregulates\b",                   "regulates",      "regulates"),
    (r"\brelat\b",                       "relat",          "relate"),
    (r"\brelatd\b",                      "relatd",         "related"),
    (r"\brelates\b",                     "relates",        "relates"),
    (r"\brelativ\b",                     "relativ",        "relative"),
    (r"\brelatives\b",                   "relatives",      "relatives"),
    (r"\breleas\b",                      "releas",         "release"),
    (r"\breleasd\b",                     "releasd",        "released"),
    (r"\breleases\b",                    "releases",       "releases"),
    (r"\brelev\b",                       "relev",          "relevant"),
    (r"\brelevants\b",                   "relevants",      "relevants"),
    (r"\breli\b",                        "reli",           "rely"),
    (r"\brelied\b",                      "relied",         "relied"),
    (r"\breli es\b",                     "reli es",        "relies"),
    (r"\bremain\b",                      "remain",         "remain"),
    (r"\bremains\b",                     "remains",        "remains"),
    (r"\bremov\b",                       "remov",          "remove"),
    (r"\bremovd\b",                      "removd",         "removed"),
    (r"\bremoves\b",                     "removes",        "removes"),
    (r"\brend\b",                        "rend",           "render"),
    (r"\brends\b",                       "rends",          "renders"),
    (r"\brenew\b",                       "renew",          "renew"),
    (r"\brenews\b",                      "renews",         "renews"),
    (r"\brepair\b",                      "repair",         "repair"),
    (r"\brepairs\b",                     "repairs",        "repairs"),
    (r"\brepeat\b",                      "repeat",         "repeat"),
    (r"\brepeats\b",                     "repeats",        "repeats"),
    (r"\breplac\b",                      "replac",         "replace"),
    (r"\breplacd\b",                     "replacd",        "replaced"),
    (r"\breplaces\b",                    "replaces",       "replaces"),
    (r"\breply\b",                       "reply",          "reply"),
    (r"\breplies\b",                     "replies",        "replies"),
    (r"\breport\b",                      "report",         "report"),
    (r"\breports\b",                     "reports",        "reports"),
    (r"\brepresent\b",                   "represent",      "represent"),
    (r"\brepresents\b",                  "represents",     "represents"),
    (r"\brequest\b",                     "request",        "request"),
    (r"\brequests\b",                    "requests",       "requests"),
    (r"\brequir\b",                      "requir",         "require"),
    (r"\brequird\b",                     "requird",        "required"),
    (r"\brequires\b",                    "requires",       "requires"),
    (r"\breserv\b",                      "reserv",         "reserve"),
    (r"\breservd\b",                     "reservd",        "reserved"),
    (r"\breserves\b",                    "reserves",       "reserves"),
    (r"\bresolv\b",                      "resolv",         "resolve"),
    (r"\bresolvd\b",                     "resolvd",        "resolved"),
    (r"\bresolves\b",                    "resolves",       "resolves"),
    (r"\bresourc\b",                     "resourc",        "resource"),
    (r"\bresources\b",                   "resources",      "resources"),
    (r"\brespons\b",                     "respons",        "respond"),
    (r"\bresponsd\b",                    "responsd",       "responded"),
    (r"\bresponses\b",                   "responses",      "responds"),
    (r"\brestor\b",                      "restor",         "restore"),
    (r"\brestord\b",                     "restord",        "restored"),
    (r"\brestores\b",                    "restores",       "restores"),
    (r"\brestrict\b",                    "restrict",       "restrict"),
    (r"\brestricts\b",                   "restricts",      "restricts"),
    (r"\bresult\b",                      "result",         "result"),
    (r"\bresults\b",                     "results",        "results"),
    (r"\bretail\b",                      "retail",         "retail"),
    (r"\bretains\b",                     "retains",        "retains"),
    (r"\bretir\b",                       "retir",          "retire"),
    (r"\bretird\b",                      "retird",         "retired"),
    (r"\bretires\b",                     "retires",        "retires"),
    (r"\breturn\b",                      "return",         "return"),
    (r"\breturns\b",                     "returns",        "returns"),
    (r"\brev\b",                         "rev",            "review"),
    (r"\breviews\b",                     "reviews",        "reviews"),
    (r"\brevis\b",                       "revis",          "revise"),
    (r"\brevisd\b",                      "revisd",         "revised"),
    (r"\brevises\b",                     "revises",        "revises"),
    (r"\brisk\b",                        "risk",           "risk"),
    (r"\brisks\b",                       "risks",          "risks"),
    (r"\bro\b",                          "ro",             "role"),
    (r"\broles\b",                       "roles",          "roles"),
    (r"\brul\b",                         "rul",            "rule"),
    (r"\brules\b",                       "rules",          "rules"),
    (r"\brun\b",                         "run",            "run"),
    (r"\bruns\b",                        "runs",           "runs"),
    (r"\bsaf\b",                         "saf",            "safe"),
    (r"\bsafes\b",                       "safes",          "safes"),
    (r"\bsatisf\b",                      "satisf",         "satisfy"),
    (r"\bsatisfid\b",                    "satisfid",       "satisfied"),
    (r"\bsatisfies\b",                   "satisfies",      "satisfies"),
    (r"\bsav\b",                         "sav",            "save"),
    (r"\bsaves\b",                       "saves",          "saves"),
    (r"\bsay\b",                         "say",            "say"),
    (r"\bsays\b",                        "says",           "says"),
    (r"\bscal\b",                        "scal",           "scale"),
    (r"\bscales\b",                      "scales",         "scales"),
    (r"\bsearch\b",                      "search",         "search"),
    (r"\bsearches\b",                    "searches",       "searches"),
    (r"\bseason\b",                      "season",         "season"),
    (r"\bseasons\b",                     "seasons",        "seasons"),
    (r"\bsecond\b",                      "second",         "second"),
    (r"\bseconds\b",                     "seconds",        "seconds"),
    (r"\bsecur\b",                       "secur",          "secure"),
    (r"\bsecurd\b",                      "securd",         "secured"),
    (r"\bsecures\b",                     "secures",        "secures"),
    (r"\bsee\b",                         "see",            "see"),
    (r"\bsees\b",                        "sees",           "sees"),
    (r"\bseek\b",                        "seek",           "seek"),
    (r"\bseeks\b",                       "seeks",          "seeks"),
    (r"\bseem\b",                        "seem",           "seem"),
    (r"\bseems\b",                       "seems",          "seems"),
    (r"\bsegment\b",                     "segment",        "segment"),
    (r"\bsegments\b",                    "segments",       "segments"),
    (r"\bselect\b",                      "select",         "select"),
    (r"\bselects\b",                     "selects",        "selects"),
    (r"\bsell\b",                        "sell",           "sell"),
    (r"\bsells\b",                       "sells",          "sells"),
    (r"\bsend\b",                        "send",           "send"),
    (r"\bsends\b",                       "sends",          "sends"),
    (r"\bsepar\b",                       "separ",          "separate"),
    (r"\bseparatd\b",                    "separatd",       "separated"),
    (r"\bseparates\b",                   "separates",      "separates"),
    (r"\bsequenc\b",                     "sequenc",        "sequence"),
    (r"\bsequences\b",                   "sequences",      "sequences"),
    (r"\bseri\b",                        "seri",           "series"),
    (r"\bseries\b",                      "series",         "series"),
    (r"\bserv\b",                        "serv",           "serve"),
    (r"\bservd\b",                       "servd",          "served"),
    (r"\bserves\b",                      "serves",         "serves"),
    (r"\bset\b",                         "set",            "set"),
    (r"\bsets\b",                        "sets",           "sets"),
    (r"\bsett\b",                        "sett",           "settle"),
    (r"\bsettd\b",                       "settd",          "settled"),
    (r"\bsettles\b",                     "settles",        "settles"),
    (r"\bsever\b",                       "sever",          "severe"),
    (r"\bsevers\b",                      "severs",         "severs"),
    (r"\bshare\b",                       "share",          "share"),
    (r"\bshares\b",                      "shares",         "shares"),
    (r"\bshift\b",                       "shift",          "shift"),
    (r"\bshifts\b",                      "shifts",         "shifts"),
    (r"\bship\b",                        "ship",           "ship"),
    (r"\bships\b",                       "ships",          "ships"),
    (r"\bshort\b",                       "short",          "short"),
    (r"\bshorts\b",                      "shorts",         "shorts"),
    (r"\bshould\b",                      "should",         "should"),
    (r"\bshoulds\b",                     "shoulds",        "shoulds"),
    (r"\bshow\b",                        "show",           "show"),
    (r"\bshows\b",                       "shows",          "shows"),
    (r"\bsign\b",                        "sign",           "sign"),
    (r"\bsigns\b",                       "signs",          "signs"),
    (r"\bsimilar\b",                     "similar",        "similar"),
    (r"\bsimilars\b",                    "similars",       "similars"),
    (r"\bsimpl\b",                       "simpl",          "simple"),
    (r"\bsimples\b",                     "simples",        "simples"),
    (r"\bsimul\b",                       "simul",          "simulate"),
    (r"\bsimulatd\b",                    "simulatd",       "simulated"),
    (r"\bsimulates\b",                   "simulates",      "simulates"),
    (r"\bsingl\b",                       "singl",          "single"),
    (r"\bsingles\b",                     "singles",        "singles"),
    (r"\bsit\b",                         "sit",            "sit"),
    (r"\bsits\b",                        "sits",           "sits"),
    (r"\bsituat\b",                      "situat",         "situate"),
    (r"\bsituatd\b",                     "situatd",        "situated"),
    (r"\bsituates\b",                    "situates",       "situates"),
    (r"\bsize\b",                        "size",           "size"),
    (r"\bsizes\b",                       "sizes",          "sizes"),
    (r"\bslow\b",                        "slow",           "slow"),
    (r"\bslows\b",                       "slows",          "slows"),
    (r"\bsmall\b",                       "small",          "small"),
    (r"\bsmalls\b",                      "smalls",         "smalls"),
    (r"\bsocial\b",                      "social",         "social"),
    (r"\bsocials\b",                     "socials",        "socials"),
    (r"\bsolv\b",                        "solv",           "solve"),
    (r"\bsolvd\b",                       "solvd",          "solved"),
    (r"\bsolves\b",                      "solves",         "solves"),
    (r"\bsort\b",                        "sort",           "sort"),
    (r"\bsorts\b",                       "sorts",          "sorts"),
    (r"\bsourc\b",                       "sourc",          "source"),
    (r"\bsources\b",                     "sources",        "sources"),
    (r"\bspecial\b",                     "special",        "special"),
    (r"\bspecials\b",                    "specials",       "specials"),
    (r"\bspecif\b",                      "specif",         "specify"),
    (r"\bspecifd\b",                     "specifd",        "specified"),
    (r"\bspecifies\b",                   "specifies",      "specifies"),
    (r"\bspend\b",                       "spend",          "spend"),
    (r"\bspends\b",                      "spends",         "spends"),
    (r"\bspok\b",                        "spok",           "spoke"),
    (r"\bspokes\b",                      "spokes",         "spokes"),
    (r"\bspread\b",                      "spread",         "spread"),
    (r"\bspreads\b",                     "spreads",        "spreads"),
    (r"\bstandard\b",                    "standard",       "standard"),
    (r"\bstandards\b",                   "standards",      "standards"),
    (r"\bstart\b",                       "start",          "start"),
    (r"\bstarts\b",                      "starts",         "starts"),
    (r"\bstate\b",                       "state",          "state"),
    (r"\bstates\b",                      "states",         "states"),
    (r"\bstat\b",                        "stat",           "status"),
    (r"\bstatuses\b",                    "statuses",       "statuses"),
    (r"\bstay\b",                        "stay",           "stay"),
    (r"\bstays\b",                       "stays",          "stays"),
    (r"\bstep\b",                        "step",           "step"),
    (r"\bsteps\b",                       "steps",          "steps"),
    (r"\bstimul\b",                      "stimul",         "stimulate"),
    (r"\bstimulatd\b",                   "stimulatd",      "stimulated"),
    (r"\bstimulates\b",                  "stimulates",     "stimulates"),
    (r"\bstop\b",                        "stop",           "stop"),
    (r"\bstops\b",                       "stops",          "stops"),
    (r"\bstore\b",                       "store",          "store"),
    (r"\bstores\b",                      "stores",         "stores"),
    (r"\bstrateg\b",                     "strateg",        "strategy"),
    (r"\bstrategies\b",                  "strategies",     "strategies"),
    (r"\bstrength\b",                    "strength",       "strength"),
    (r"\bstrengths\b",                   "strengths",       "strengths"),
    (r"\bstress\b",                      "stress",         "stress"),
    (r"\bstresses\b",                    "stresses",       "stresses"),
    (r"\bstructur\b",                    "structur",       "structure"),
    (r"\bstructures\b",                  "structures",     "structures"),
    (r"\bstruggl\b",                     "struggl",        "struggle"),
    (r"\bstruggld\b",                    "struggld",       "struggled"),
    (r"\bstruggles\b",                   "struggles",      "struggles"),
    (r"\bstudi\b",                       "studi",          "study"),
    (r"\bstudid\b",                      "studid",         "studied"),
    (r"\bstudies\b",                     "studies",        "studies"),
    (r"\bstuff\b",                       "stuff",          "stuff"),
    (r"\bstuffs\b",                      "stuffs",         "stuffs"),
    (r"\bsubmit\b",                      "submit",         "submit"),
    (r"\bsubmits\b",                     "submits",        "submits"),
    (r"\bsucceed\b",                     "succeed",        "succeed"),
    (r"\bsucceeds\b",                    "succeeds",       "succeeds"),
    (r"\bsuffer\b",                      "suffer",         "suffer"),
    (r"\bsuffers\b",                     "suffers",        "suffers"),
    (r"\bsuggest\b",                     "suggest",        "suggest"),
    (r"\bsuggests\b",                    "suggests",       "suggests"),
    (r"\bsuit\b",                        "suit",           "suit"),
    (r"\bsuits\b",                       "suits",          "suits"),
    (r"\bsummar\b",                      "summar",         "summarize"),
    (r"\bsummarizd\b",                   "summarizd",      "summarized"),
    (r"\bsummarizes\b",                  "summarizes",     "summarizes"),
    (r"\bsupervis\b",                    "supervis",       "supervise"),
    (r"\bsupervisd\b",                   "supervisd",      "supervised"),
    (r"\bsupervises\b",                  "supervises",     "supervises"),
    (r"\bsuppli\b",                      "suppli",         "supply"),
    (r"\bsupplid\b",                     "supplid",        "supplied"),
    (r"\bsupplies\b",                    "supplies",       "supplies"),
    (r"\bsupport\b",                     "support",        "support"),
    (r"\bsupports\b",                    "supports",       "supports"),
    (r"\bsurfac\b",                      "surfac",         "surface"),
    (r"\bsurfaces\b",                    "surfaces",       "surfaces"),
    (r"\bsurpass\b",                     "surpass",        "surpass"),
    (r"\bsurpasses\b",                   "surpasses",      "surpasses"),
    (r"\bsurvey\b",                      "survey",         "survey"),
    (r"\bsurveys\b",                     "surveys",        "surveys"),
    (r"\bsustain\b",                     "sustain",        "sustain"),
    (r"\bsustains\b",                    "sustains",       "sustains"),
    (r"\bswitch\b",                      "switch",         "switch"),
    (r"\bswitches\b",                    "switches",       "switches"),
    (r"\bsymbol\b",                      "symbol",         "symbol"),
    (r"\bsymbols\b",                     "symbols",        "symbols"),
    (r"\bsystem\b",                      "system",         "system"),
    (r"\bsystems\b",                     "systems",        "systems"),
    (r"\btabl\b",                        "tabl",           "table"),
    (r"\btables\b",                      "tables",         "tables"),
    (r"\btak\b",                         "tak",            "take"),
    (r"\btakes\b",                       "takes",          "takes"),
    (r"\btalk\b",                        "talk",           "talk"),
    (r"\btalks\b",                       "talks",          "talks"),
    (r"\btarg\b",                        "targ",           "target"),
    (r"\btargets\b",                     "targets",        "targets"),
    (r"\btask\b",                        "task",           "task"),
    (r"\btasks\b",                       "tasks",          "tasks"),
    (r"\bteach\b",                       "teach",          "teach"),
    (r"\bteaches\b",                     "teaches",        "teaches"),
    (r"\bteam\b",                        "team",           "team"),
    (r"\bteams\b",                       "teams",          "teams"),
    (r"\btechnic\b",                     "technic",        "technical"),
    (r"\btechnicals\b",                  "technicals",     "technicals"),
    (r"\btechnolog\b",                   "technolog",      "technology"),
    (r"\btechnologies\b",                "technologies",   "technologies"),
    (r"\btell\b",                        "tell",           "tell"),
    (r"\btells\b",                       "tells",          "tells"),
    (r"\bterm\b",                        "term",           "term"),
    (r"\bterms\b",                       "terms",          "terms"),
    (r"\btest\b",                        "test",           "test"),
    (r"\btests\b",                       "tests",          "tests"),
    (r"\btext\b",                        "text",           "text"),
    (r"\btexts\b",                       "texts",          "texts"),
    (r"\bthank\b",                       "thank",          "thank"),
    (r"\bthanks\b",                      "thanks",         "thanks"),
    (r"\btheor\b",                       "theor",          "theory"),
    (r"\btheories\b",                    "theories",       "theories"),
    (r"\bthink\b",                       "think",          "think"),
    (r"\bthinks\b",                      "thinks",          "thinks"),
    (r"\bthreat\b",                      "threat",         "threat"),
    (r"\bthreats\b",                     "threats",        "threats"),
    (r"\bthrough\b",                     "through",        "through"),
    (r"\bthroughs\b",                    "throughs",       "throughs"),
    (r"\btime\b",                        "time",           "time"),
    (r"\btimes\b",                       "times",          "times"),
    (r"\btitl\b",                        "titl",           "title"),
    (r"\btitles\b",                      "titles",         "titles"),
    (r"\btool\b",                        "tool",           "tool"),
    (r"\btools\b",                       "tools",          "tools"),
    (r"\btotal\b",                       "total",          "total"),
    (r"\btotals\b",                      "totals",         "totals"),
    (r"\btrack\b",                       "track",          "track"),
    (r"\btracks\b",                      "tracks",         "tracks"),
    (r"\btrad\b",                        "trad",           "trade"),
    (r"\btrades\b",                      "trades",         "trades"),
    (r"\btrain\b",                       "train",          "train"),
    (r"\btrains\b",                      "trains",          "trains"),
    (r"\btransact\b",                    "transact",       "transact"),
    (r"\btransacts\b",                   "transacts",      "transacts"),
    (r"\btransfer\b",                    "transfer",       "transfer"),
    (r"\btransfers\b",                   "transfers",      "transfers"),
    (r"\btransform\b",                   "transform",      "transform"),
    (r"\btransforms\b",                  "transforms",     "transforms"),
    (r"\btransit\b",                     "transit",        "transit"),
    (r"\btransits\b",                    "transits",       "transits"),
    (r"\btranslat\b",                    "translat",       "translate"),
    (r"\btranslatd\b",                   "translatd",      "translated"),
    (r"\btranslates\b",                  "translates",     "translates"),
    (r"\btransmit\b",                    "transmit",       "transmit"),
    (r"\btransmits\b",                   "transmits",      "transmits"),
    (r"\btransport\b",                   "transport",      "transport"),
    (r"\btransports\b",                  "transports",     "transports"),
    (r"\btreat\b",                       "treat",          "treat"),
    (r"\btreats\b",                      "treats",         "treats"),
    (r"\btrend\b",                       "trend",          "trend"),
    (r"\btrends\b",                      "trends",         "trends"),
    (r"\btri\b",                         "tri",            "try"),
    (r"\btries\b",                       "tries",          "tries"),
    (r"\btrigger\b",                     "trigger",        "trigger"),
    (r"\btriggers\b",                    "triggers",       "triggers"),
    (r"\bturn\b",                        "turn",           "turn"),
    (r"\bturns\b",                       "turns",          "turns"),
    (r"\btype\b",                        "type",           "type"),
    (r"\btypes\b",                       "types",          "types"),
    (r"\bunderstand\b",                  "understand",     "understand"),
    (r"\bunderstands\b",                 "understands",    "understands"),
    (r"\bunit\b",                        "unit",           "unit"),
    (r"\bunits\b",                       "units",          "units"),
    (r"\bunivers\b",                     "univers",        "universal"),
    (r"\buniversals\b",                  "universals",     "universals"),
    (r"\bupdat\b",                       "updat",          "update"),
    (r"\bupdatd\b",                      "updatd",         "updated"),
    (r"\bupdates\b",                     "updates",        "updates"),
    (r"\bus\b",                          "us",             "use"),
    (r"\buses\b",                        "uses",           "uses"),
    (r"\butil\b",                        "util",           "utility"),
    (r"\butilities\b",                   "utilities",      "utilities"),
    (r"\bvalid\b",                       "valid",          "valid"),
    (r"\bvalids\b",                      "valids",         "valids"),
    (r"\bvalu\b",                        "valu",           "value"),
    (r"\bvalues\b",                      "values",         "values"),
    (r"\bvariat\b",                      "variat",         "vary"),
    (r"\bvaryd\b",                       "varyd",          "varied"),
    (r"\bvaries\b",                      "varies",         "varies"),
    (r"\bvari\b",                        "vari",           "vary"),
    (r"\bvaryd\b",                       "varyd",          "varied"),
    (r"\bvaries\b",                      "varies",         "varies"),
    (r"\bvehicl\b",                      "vehicl",         "vehicle"),
    (r"\bvehicles\b",                    "vehicles",       "vehicles"),
    (r"\bversion\b",                     "version",        "version"),
    (r"\bversions\b",                    "versions",       "versions"),
    (r"\bview\b",                        "view",           "view"),
    (r"\bviews\b",                       "views",          "views"),
    (r"\bvisit\b",                       "visit",          "visit"),
    (r"\bvisits\b",                      "visits",         "visits"),
    (r"\bvisual\b",                      "visual",         "visual"),
    (r"\bvisuals\b",                     "visuals",        "visuals"),
    (r"\bvolum\b",                       "volum",          "volume"),
    (r"\bvolumes\b",                     "volumes",        "volumes"),
    (r"\bwait\b",                        "wait",           "wait"),
    (r"\bwaits\b",                       "waits",          "waits"),
    (r"\bwant\b",                        "want",           "want"),
    (r"\bwants\b",                       "wants",          "wants"),
    (r"\bwarn\b",                        "warn",           "warn"),
    (r"\bwarns\b",                       "warns",          "warns"),
    (r"\bwarrant\b",                     "warrant",        "warrant"),
    (r"\bwarrants\b",                    "warrants",       "warrants"),
    (r"\bwatch\b",                       "watch",          "watch"),
    (r"\bwatches\b",                     "watches",        "watches"),
    (r"\bway\b",                         "way",            "way"),
    (r"\bways\b",                        "ways",           "ways"),
    (r"\bweak\b",                        "weak",           "weak"),
    (r"\bweaks\b",                       "weaks",          "weaks"),
    (r"\bwear\b",                        "wear",           "wear"),
    (r"\bwears\b",                       "wears",          "wears"),
    (r"\bweek\b",                        "week",           "week"),
    (r"\bweeks\b",                       "weeks",          "weeks"),
    (r"\bweight\b",                      "weight",         "weight"),
    (r"\bweights\b",                     "weights",        "weights"),
    (r"\bwell\b",                        "well",           "well"),
    (r"\bwells\b",                       "wells",          "wells"),
    (r"\bwill\b",                        "will",           "will"),
    (r"\bwills\b",                       "wills",          "wills"),
    (r"\bwin\b",                         "win",            "win"),
    (r"\bwins\b",                        "wins",           "wins"),
    (r"\bwithdraw\b",                    "withdraw",       "withdraw"),
    (r"\bwithdraws\b",                   "withdraws",      "withdraws"),
    (r"\bwithhold\b",                    "withhold",       "withhold"),
    (r"\bwithholds\b",                   "withholds",       "withholds"),
    (r"\bwithin\b",                      "within",         "within"),
    (r"\bwithins\b",                     "withins",        "withins"),
    (r"\bwithout\b",                     "without",        "without"),
    (r"\bwithouts\b",                    "withouts",       "withouts"),
    (r"\bword\b",                        "word",           "word"),
    (r"\bwords\b",                       "words",          "words"),
    (r"\bwork\b",                        "work",           "work"),
    (r"\bworks\b",                       "works",          "works"),
    (r"\bworld\b",                       "world",          "world"),
    (r"\bworlds\b",                      "worlds",          "worlds"),
    (r"\bworth\b",                       "worth",          "worth"),
    (r"\bworths\b",                      "worths",         "worths"),
    (r"\bwould\b",                       "would",          "would"),
    (r"\bwoulds\b",                      "woulds",         "woulds"),
    (r"\bwrite\b",                       "write",          "write"),
    (r"\bwrites\b",                      "writes",         "writes"),
    (r"\byear\b",                        "year",           "year"),
    (r"\byears\b",                       "years",          "years"),
    (r"\byield\b",                       "yield",          "yield"),
    (r"\byields\b",                      "yields",         "yields"),
    (r"\bzero\b",                        "zero",           "zero"),
    (r"\bzeros\b",                       "zeros",          "zeros"),
    (r"\bzone\b",                        "zone",           "zone"),
    (r"\bzones\b",                       "zones",          "zones"),
    # --- 마스터 데이터 분석 추가 패턴 (2026-03) ---
    # 오탈자: l/t/f 누락/전치 (마스터 데이터 직접 발견)
    (r"\bliabilties\b",              "liabilties",    "liabilities"),      # t 누락, liabilites와 별개
    (r"\bacqusition\b",              "acqusition",    "acquisition"),      # i 누락
    (r"\binterst\b",                 "interst",       "interest"),         # e 누락
    (r"\bgurantee\b",                "gurantee",      "guarantee"),
    (r"\bgurantees\b",               "gurantees",     "guarantees"),
    (r"\bmeasurment\b",              "measurment",    "measurement"),
    (r"\bpreiod\b",                  "preiod",        "period"),
    (r"\bprovison\b",                "provison",      "provision"),
    (r"\bprovisons\b",               "provisons",     "provisions"),
    (r"\bdefered\b",                 "defered",       "deferred"),
    (r"\brecieved\b",                "recieved",      "received"),
    (r"\brrofit\b",                  "rrofit",        "profit"),           # 앞 r 중복
    (r"\bpropert\b",                 "propert",       "property"),         # 끝 y 누락
    (r"\bteminal\b",                 "teminal",       "terminal"),
    (r"\bdisacounts\b",              "disacounts",    "discounts"),
    (r"\brealated\b",                "realated",      "related"),
    (r"\bofrevenue\b",               "ofrevenue",     "of revenue"),       # 공백 누락
    (r"\bborrowingsl\b",             "borrowingsl",   "borrowings"),       # 끝에 l 오붙음
    (r"\boutlows?\b",                "outlow(s)",     "outflow(s)"),       # f 누락 (6건)
    (r"(?<!\w)ncome\b",              "ncome",         "Income"),           # 문장 첫 I 누락 (3건)
    (r"\biDescription\b",           "iDescription",  "Description"),      # 앞 i 오붙음
    # 오탈자: 예방적 추가 (재무 영문명 흔한 오탈자)
    (r"\bforiegn\b",                 "foriegn",       "foreign"),
    (r"\boccured\b",                 "occured",       "occurred"),
    (r"\btransfered\b",              "transfered",    "transferred"),
    (r"\bseperate\b",                "seperate",      "separate"),
    (r"\bseperately\b",              "seperately",    "separately"),
    (r"\breconcilation\b",           "reconcilation", "reconciliation"),
    (r"\bclasification\b",           "clasification", "classification"),
    (r"\bstatment\b",                "statment",      "statement"),
    (r"\bsettlment\b",               "settlment",     "settlement"),
    (r"\bcontibution\b",             "contibution",   "contribution"),
    (r"\bcontibutions\b",            "contibutions",  "contributions"),
    (r"\bvalution\b",                "valution",      "valuation"),
    (r"\bvaulation\b",               "vaulation",     "valuation"),
    (r"\bassest\b",                  "assest",        "assets"),
    (r"\bpayement\b",                "payement",      "payment"),
    (r"\bpaymnet\b",                 "paymnet",       "payment"),
    # 구조 오류
    (r"\b\w+\?\w",                   "word?word",     "word-word (하이픈→? 변환 오류)"),   # 365건
    (r",,",                          ",,",            ", (이중 쉼표)"),                     # 7건
    (r"\(\s*\)",                     "()",            "빈 괄호 — 내용 채우거나 제거"),       # 3건
    # --- 이중 공백은 별도 체크 ---
]

# 대/소문자 오류: en_label 첫 글자는 대문자여야 하는 경우 체크는 
# 자연어 문장형 레이블에만 적용 (CamelCase가 아닌 경우)

# ─────────────────────────────────────────────────────────────
# Known XBRL suffix patterns (these indicate XBRL IDs, not labels)
# ─────────────────────────────────────────────────────────────
XBRL_SUFFIX_RE = re.compile(
    r"(Abstract|Table|LineItems|Lineitems|LineItem|Member|Axis|Domain|"
    r"Explanatory|Explantion|LinkRole|Roll[Ff]orward)$"
)

# XBRL 구조 요소 타입 (Abstract, Table 등은 en_label이 XBRL ID여도 허용)
STRUCTURAL_SUFFIX_RE = re.compile(
    r"\[(abstract|table|line\s*items|항목|개요|표|문장영역)\]",
    re.IGNORECASE,
)

# 회사 고유 확장 Member 이름: CamelCase + Member 로 끝나는 패턴
ENTITY_MEMBER_RE = re.compile(r"^[A-Z][A-Za-z]+Member$")

# ─────────────────────────────────────────────────────────────
# Core detection logic
# ─────────────────────────────────────────────────────────────
def _is_camel_case_id(s: str) -> bool:
    """True if s looks like a pure XBRL CamelCase element ID (no spaces, ≥2 uppercase transitions)."""
    if " " in s:
        return False
    # At least 2 capital letter transitions (Word boundaries)
    caps = len(re.findall(r"[A-Z]", s))
    return caps >= 3 and bool(re.match(r"^[A-Z][a-z]", s))


def _contains_korean(s: str) -> bool:
    return bool(re.search(r"[\uAC00-\uD7A3\u1100-\u11FF\u3130-\u318F]", s))


def _detect_typos_in(text: str, field: str = "en") -> list[Issue]:
    """
    단순 오탈자를 주어진 텍스트에서 검출한다.
    field: "en"  → highlight_en  (기본 영문명)
           "en2" → highlight_en2 (표현 영문명)
    """
    issues: list[Issue] = []
    for pattern, typo, correct in TYPO_RULES:
        if typo is None:
            continue
        if re.search(pattern, text, re.IGNORECASE):
            kw = {"highlight_" + field: [typo]}
            issues.append(Issue(
                error_type  = "단순 오탈자",
                description = f'"{typo}" → "{correct}" 로 수정 필요',
                **kw,
            ))
    if "  " in text:
        issues.append(Issue(
            error_type  = "단순 오탈자",
            description = "이중 공백 포함 — 단일 공백으로 수정 필요",
        ))
    # ── 마침표 ──────────────────────────────────────────────────
    _SAFE_DOT_ENDS = re.compile(
        r'(etc\.|Co\.|Ltd\.|Inc\.|Corp\.|S\.A\.|e\.g\.|i\.e\.|vs\.|No\.|Fig\.)$',
        re.IGNORECASE,
    )
    text_s = text.strip()
    if (
        text_s.endswith(".")
        and not text_s.endswith("...")
        and not _SAFE_DOT_ENDS.search(text_s)
    ):
        kw = {"highlight_" + field: ["."]}
        issues.append(Issue(
            error_type  = "단순 오탈자",
            description = f'{"기본" if field == "en" else "표현"} 영문명 끝에 마침표(".") 불필요',
            **kw,
        ))
    return issues


def detect_issues(
    en:          str,
    ko:          str  = "",
    en2:         str  = "",
    ko2:         str  = "",
    prefix:      str  = "",
    label_title: str  = "기본",
) -> list[Issue]:
    """
    오탈자·원칙 위배를 탐지한다.

    오탈자 검사 범위 (가이드 §4.Ⅰ.3 기반):
        - prefix == 'entity'        → 기본 영문명(en_label) 오탈자 검사
        - prefix != 'entity'
            labelTitle != '기본'    → 표현 영문명(en2)만 오탈자 검사
            labelTitle == '기본'    → 오탈자 검사 없음 (표준 계정은 검사 불필요)

    영문명 미기재 / XBRL 확장 원칙 위배는 entity prefix 항목에만 적용.
    """
    is_entity = prefix.startswith("entity") or prefix == ""
    en_s      = en.strip()
    issues: list[Issue] = []

    # ═══════════════════════════════════════════════════════════
    # entity prefix: 전체 검사 (미기재 + 확장원칙 + 오탈자)
    # ═══════════════════════════════════════════════════════════
    if is_entity:

        # ── (A) 영문명 미기재 ─────────────────────────────────
        if not en_s:
            return []

        PLACEHOLDER_RE = re.compile(
            r"^(item|title)([-_]\d+)?(\s*\[.*\])?\s*$", re.IGNORECASE
        )
        if PLACEHOLDER_RE.match(en_s):
            issues.append(Issue(
                error_type       = "영문명 미기재",
                description      = (
                    f"⚠ 정정공시 대상 ⚠ — 기본 영문명에 적절한 영문 표준명이 기재되지 않음 "
                    f'(현재값: "{en_s}")'
                ),
                highlight_en     = [en_s],
                is_full_en_error = True,
            ))
            return issues

        # ── 한글 annotation suffix 패턴: "English text [구성요소|축]" ─
        # XBRL 요소 유형 annotation이 한글로 입력된 경우
        # (en_label에서는 [member]/[axis]/[table] 등 영문을 써야 함)
        _KR_ANNOT_MAP = {
            "구성요소": "member",
            "축":       "axis",
            "표":       "table",
            "항목":     "line items",
            "개요":     "abstract",
            "문장영역": "text block",
        }
        _KR_ANNOT_RE = re.compile(
            r"\[(" + "|".join(_KR_ANNOT_MAP.keys()) + r")\]",
            re.IGNORECASE,
        )
        annot_match = _KR_ANNOT_RE.search(en_s)
        eng_without_annot = _KR_ANNOT_RE.sub("", en_s).strip()

        if annot_match and not _contains_korean(eng_without_annot):
            # 순수 영문 본문 + 한글 suffix
            kr_sfx  = annot_match.group(1)
            en_sfx  = _KR_ANNOT_MAP.get(kr_sfx, kr_sfx)
            issues.append(Issue(
                error_type       = "단순 오탈자",
                description      = (
                    f"영문명 내 XBRL 요소 유형이 한글로 입력됨 — "
                    f'"[{kr_sfx}]" → "[{en_sfx}]" 로 수정 필요'
                ),
                highlight_en     = [annot_match.group()],
                is_full_en_error = False,
            ))
            return issues

        if _contains_korean(en_s):
            issues.append(Issue(
                error_type       = "영문명 미기재",
                description      = "기본 영문명 필드에 한글이 입력됨 — 영문 표준명으로 대체 필요",
                is_full_en_error = True,
            ))
            return issues

        # ── (B) XBRL 확장 원칙 위배 ───────────────────────────
        violations: list[str] = []
        hi_ext: list[str]     = []

        # (B-1) FootNote 포함
        m_fn = re.search(r"foot\s*note", en, re.IGNORECASE)
        if m_fn:
            hi_ext.append(m_fn.group())
            violations.append(
                '"FootNote" 포함 — 주석 참조 형식은 확장 요소 기본 영문명으로 부적합 '
                "(가이드 §4.Ⅰ.3)"
            )

        # (B-2) ~주기 / Disclosure Note 형식
        if re.search(r"주기", ko + ko2):
            violations.append(
                '"~주기" 형식 — XBRL 확장 원칙상 독립 요소로 부적합 (가이드 §4.Ⅰ.3)'
            )
        if re.search(r"disclosure\s+note", en, re.IGNORECASE):
            hi_ext.append("disclosure note")
            violations.append('"Disclosure Note" 형식 — 가이드 §4.Ⅰ.3 위배')

        # (B-3) CamelCase XBRL ID
        is_structural = bool(STRUCTURAL_SUFFIX_RE.search(en_s))
        if not is_structural and _is_camel_case_id(en_s) and not XBRL_SUFFIX_RE.search(en_s):
            hi_ext.append(en_s[:80])
            violations.append(
                "CamelCase XBRL 요소 ID가 기본 영문명으로 입력됨 — "
                "공백 구분 자연어 형식(Natural Language Label)으로 기재 필요 "
                "(가이드 §3.Ⅱ.2.(4), §4.Ⅰ.3.(3))"
            )

        # (B-4) DescriptionOf...
        if re.match(r"^Description[A-Z]", en_s) and " " not in en_s[:50]:
            hi_ext.append(en_s[:60])
            violations.append(
                '"DescriptionOf..." 형태의 XBRL 요소 ID가 기본 영문명으로 입력됨 '
                "(가이드 §3.Ⅱ.2.(4))"
            )

        # (B-5) XxxAndXxx
        if re.match(r"^[A-Z][a-z]+And[A-Z]", en_s) and " " not in en_s:
            m2 = re.match(r"^[A-Za-z]+", en_s)
            hi_ext.append(m2.group() if m2 else en_s[:30])
            violations.append(
                "XBRL 요소 ID 형식(XxxAndXxx)이 기본 영문명으로 입력됨 "
                "(가이드 §3.Ⅱ.2.(4))"
            )

        # (B-6) Member 접미사 누락
        if (
            _is_camel_case_id(en_s)
            and not XBRL_SUFFIX_RE.search(en_s)
            and re.search(r"(Co|Corp|Ltd|Inc|Llc|Plc|Gmbh|SA|SAS|BV)$", en_s, re.IGNORECASE)
        ):
            hi_ext.append(en_s[-20:])
            violations.append(
                '구성요소(Member) 이름에 "Member" 접미사 누락 — '
                'XBRL 표준상 구성요소 ID 마지막에 "Member"를 붙여야 함 '
                "(가이드 §2.Ⅱ.2.(3), §5.Ⅲ.19.(3))"
            )

        # (B-7) 너무 짧은 영문명
        if re.match(r"^[A-Za-z]{1,3}$", en_s):
            violations.append(
                f'기본 영문명이 너무 짧음("{en_s}") — 회계 개념을 충분히 설명하는 영문명 기재 필요 '
                "(가이드 §3.Ⅱ.2.(4))"
            )

        # (B-9) 현금흐름 조정 접두사 누락
        if re.search(r"현금흐름.*조정|조정.*현금흐름", ko) and not re.search(
            r"AdjustmentsFor", en_s
        ):
            violations.append(
                "현금흐름표 조정 항목 — XBRL 표준상 이름(Name)을 "
                '"AdjustmentsFor"로 시작해야 함 (가이드 §4.Ⅰ.6)'
            )

        if violations:
            issues.append(Issue(
                error_type       = "XBRL 확장 원칙 위배",
                description      = " | ".join(violations),
                highlight_en     = hi_ext,
                is_full_en_error = True,
            ))

        # ── (C) 오탈자: 기본 영문명(en) ────────────────────────
        issues.extend(_detect_typos_in(en, field="en"))

    # ═══════════════════════════════════════════════════════════
    # 비(非) entity prefix: labelTitle이 '기본'이 아닌 경우만
    # 표현 영문명(en2)에 대해서만 오탈자 검사
    # ═══════════════════════════════════════════════════════════
    elif label_title != "기본":
        if en2:
            issues.extend(_detect_typos_in(en2, field="en2"))

    # prefix != entity + labelTitle == '기본' → 검사 없음

    return issues


# ─────────────────────────────────────────────────────────────
# Excel output helpers
# ─────────────────────────────────────────────────────────────
def _make_rich_text(
    text: str,
    patterns: list[str],
    bold: bool = False,
    italic: bool = False,
    base_color: str | None = None,
) -> "CellRichText | str":
    """
    패턴에 매칭되는 부분을 빨간 글자로 강조한 CellRichText를 반환한다.
    매칭이 없으면 원본 문자열 반환.

    InlineFont.sz 는 포인트 단위 (9 = 9pt).
    CellRichText 셀에 c.font 를 별도로 설정하면 크기가 재해석될 수 있으므로
    normal/red 모두 sz 를 명시적으로 고정한다.
    """
    if not text or not patterns:
        return text or ""
    normal = InlineFont(
        rFont=FONT_NAME, sz=FONT_SIZE,
        b=bold, i=italic,
        color=base_color or "000000",
    )
    red = InlineFont(
        rFont=FONT_NAME, sz=FONT_SIZE,
        b=bold, i=italic,
        color=VB_RED,
    )
    pos: list[tuple[int, int]] = []
    for p in patterns:
        for m in re.finditer(re.escape(p), text, re.IGNORECASE):
            pos.append((m.start(), m.end()))
    if not pos:
        return text
    pos.sort()
    merged: list[tuple[int, int]] = []
    for s, e in pos:
        if merged and s <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))
        else:
            merged.append((s, e))
    blocks: list[TextBlock] = []
    prev = 0
    for s, e in merged:
        if prev < s:
            blocks.append(TextBlock(normal, text[prev:s]))
        blocks.append(TextBlock(red, text[s:e]))
        prev = e
    if prev < len(text):
        blocks.append(TextBlock(normal, text[prev:]))
    return CellRichText(*blocks)


def _format_description(desc: str) -> "CellRichText | str":
    """
    ErrorDescription 셀 값을 포맷한다.

    '정정공시 대상' 포함 시:
        1행(빨강 볼드): "정정공시 대상"
        2행(일반):      나머지 설명 (이모지·구분자 제거)
    그 외: 원본 문자열 반환.
    """
    if "정정공시 대상" not in desc:
        return desc

    # 이모지 및 앞뒤 구분자 제거 후 나머지 추출
    # 예: "⚠ 정정공시 대상 ⚠ — 기본 영문명에 ..."  →  "기본 영문명에 ..."
    EMOJI_RE  = re.compile(r"[⚠️\u26A0\uFE0F\u203C\u2757]+")
    stripped  = EMOJI_RE.sub("", desc)
    # "정정공시 대상" 앞뒤를 분리
    parts = re.split(r"정정공시\s*대상", stripped, maxsplit=1)
    rest  = parts[1] if len(parts) > 1 else ""
    # 앞쪽 구분자 제거 (—, -, ─, 공백)
    rest  = re.sub(r"^[\s\-—─]+", "", rest).strip()

    # 첫 번째 줄: 빨강 (sz=포인트 단위)
    red_if = InlineFont(rFont=FONT_NAME, sz=FONT_SIZE, b=True,  color=VB_RED)
    # 두 번째 줄: 일반
    nor_if = InlineFont(rFont=FONT_NAME, sz=FONT_SIZE, color="333333")
    if rest:
        return CellRichText(
            TextBlock(red_if, "정정공시 대상"),
            TextBlock(nor_if, f"\n{rest}"),
        )
    return CellRichText(TextBlock(red_if, "정정공시 대상"))


def _apply_full_border(ws, nrows: int, ncols: int) -> None:
    inner = Side(style="dotted", color="AAAAAA")
    outer = Side(style="thin",   color="000000")
    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = Border(
                left   = outer if c == 1     else inner,
                right  = outer if c == ncols else inner,
                top    = outer if r == 1     else inner,
                bottom = outer if r == nrows else inner,
            )


# ─────────────────────────────────────────────────────────────
# Report generation
HEADERS    = [
    "Index", "Report",
    "기본 한글명", "기본 영문명",
    "표현속성", "표현 한글명", "표현 영문명",
    "ErrorType", "ErrorDescription",
]
COL_WIDTHS = [7, 28, 32, 55, 18, 30, 55, 22, 70]

# 마스터 파일용 헤더
MASTER_HEADERS = [
    "Index", "Title",
    "한글명", "영문명",
    "오류유형", "오류내용",
]
MASTER_COL_WIDTHS = [7, 28, 32, 55, 22, 70]


def generate_master_report(
    results: list[RowResult],
    out_path,  # Path | io.BytesIO
    total_rows: int,
) -> None:
    """
    마스터 파일 검사 결과를 Excel 파일로 출력한다.
    마스터 파일은 표현 필드가 없고 간단한 구조를 가진다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "마스터검사결과"

    COLS = {h: i + 1 for i, h in enumerate(MASTER_HEADERS)}
    N = len(MASTER_HEADERS)

    # ── 공통 스타일 ────────────────────────────────────────────
    hdr_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color=HEADER_FG)
    hdr_fill = PatternFill("solid", start_color=HEADER_BG)
    norm_font = Font(name=FONT_NAME, size=FONT_SIZE)
    err_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color="CC0000")
    desc_font = Font(name=FONT_NAME, italic=True, size=FONT_SIZE, color="333333")
    yell_fill = PatternFill("solid", start_color=VB_YELLOW)
    TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── 헤더 행 ────────────────────────────────────────────────
    for ci, (h, w) in enumerate(zip(MASTER_HEADERS, MASTER_COL_WIDTHS), 1):
        cell = ws.cell(1, ci, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = TOP
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # ── 데이터 없음 ────────────────────────────────────────────
    if not results:
        c = ws.cell(2, 1, "이상 없음")
        c.font = Font(name=FONT_NAME, bold=True, size=10, color="008000")
        c.alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells("A2:F2")
        _apply_full_border(ws, 2, N)
        ws.freeze_panes = "A2"
        wb.save(out_path)
        return

    # ── RowResult → 플랫 행 목록 ───────────────────────────────
    flat: list[dict] = []
    row_idx = 1  # Excel 행 번호 (헤더가 1행이므로 데이터는 2행부터)

    for rr in results:
        if not rr.issues:
            continue  # 마스터 파일은 오류가 있는 행만 출력

        for issue in rr.issues:
            row_idx += 1
            flat.append({
                "Index": row_idx - 1,  # 1부터 시작하는 인덱스
                "Title": rr.title,
                "한글명": rr.ko,
                "영문명": rr.en,
                "오류유형": issue.error_type,
                "오류내용": issue.description,
            })

    # ── 데이터 행 ───────────────────────────────────────────────
    for ri, row_data in enumerate(flat, 2):  # 2행부터 시작
        # 기본 값들
        ws.cell(ri, COLS["Index"], row_data["Index"])
        ws.cell(ri, COLS["Title"], row_data["Title"])
        ws.cell(ri, COLS["한글명"], row_data["한글명"])
        ws.cell(ri, COLS["영문명"], row_data["영문명"])
        ws.cell(ri, COLS["오류유형"], row_data["오류유형"])
        ws.cell(ri, COLS["오류내용"], row_data["오류내용"])

        # 스타일 적용
        for ci in range(1, N + 1):
            cell = ws.cell(ri, ci)
            cell.font = norm_font
            cell.alignment = TOP
            _apply_full_border(ws, ri, N)

    ws.freeze_panes = "A2"
    wb.save(out_path)


def generate_report(
    results:    list[RowResult],
    out_path,           # Path | io.BytesIO
    company:    str,
    btype:      str,
    total_rows: int,
) -> None:
    """
    검사 결과를 Excel 파일로 출력한다.

    변경 사항:
    - 헤더 배경 DEDEDE (연한 회색), 글자 진한 색
    - [AI] 접두사 제거 (ErrorType 표시 시)
    - 정정공시 대상: 이모지 제거, 1행(빨강) + 줄바꿈 + 2행(일반)
    - 표현 한글명 / 표현 영문명에도 오류 시 interior·font 색상 적용
    - 전체 데이터 ListObject (Excel Table) 처리
    - 가운데 맞춤 없음, 위쪽 맞춤 + 줄바꿈
    - insert_rows 제거 → 플랫(flat) 행 순서로 직접 기재
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "검사결과"

    COLS = {h: i + 1 for i, h in enumerate(HEADERS)}
    N    = len(HEADERS)

    # ── 공통 스타일 ────────────────────────────────────────────
    hdr_font  = Font(name=FONT_NAME, bold=True,  size=FONT_SIZE, color=HEADER_FG)
    hdr_fill  = PatternFill("solid", start_color=HEADER_BG)
    norm_font = Font(name=FONT_NAME, size=FONT_SIZE)
    err_font  = Font(name=FONT_NAME, bold=True,  size=FONT_SIZE, color="CC0000")
    desc_font = Font(name=FONT_NAME, italic=True, size=FONT_SIZE, color="333333")
    yell_fill = PatternFill("solid", start_color=VB_YELLOW)
    TOP       = Alignment(horizontal="left", vertical="top", wrap_text=True)
    TOP_CTR   = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── 헤더 행 ────────────────────────────────────────────────
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell           = ws.cell(1, ci, h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = TOP
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # ── 데이터 없음 ────────────────────────────────────────────
    if not results:
        c = ws.cell(2, 1, "이상 없음")
        c.font      = Font(name=FONT_NAME, bold=True, size=10, color="008000")
        c.alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells("A2:I2")
        _apply_full_border(ws, 2, N)
        ws.freeze_panes = "A2"
        wb.save(out_path)
        return

    # ── RowResult → 플랫 행 목록 ───────────────────────────────
    # 각 RowResult의 오류 판정을 필드별로 독립적으로 계산한 후 평탄화한다.
    flat: list[dict] = []

    for rr in results:
        all_hi_en  = [p for iss in rr.issues for p in iss.highlight_en]
        all_hi_ko  = [p for iss in rr.issues for p in iss.highlight_ko]
        all_hi_en2 = [p for iss in rr.issues for p in iss.highlight_en2]
        all_hi_ko2 = [p for iss in rr.issues for p in iss.highlight_ko2]

        is_full_en  = any(i.is_full_en_error  for i in rr.issues)
        is_full_en2 = any(i.is_full_en2_error for i in rr.issues)

        def _has_match(patterns, text):
            return bool(patterns and any(
                re.search(re.escape(p), text or "", re.I) for p in patterns
            ))

        # 필드별 오류 여부 (interior.color 적용 기준)
        en_err  = is_full_en  or _has_match(all_hi_en,  rr.en)
        ko_err  = _has_match(all_hi_ko,  rr.ko)
        en2_err = is_full_en2 or (bool(rr.en2) and _has_match(all_hi_en2, rr.en2))
        ko2_err = bool(rr.ko2) and _has_match(all_hi_ko2, rr.ko2)

        # 오탈자 패턴 강조 포함 rich-text 값 (매칭 없으면 원본 str 반환)
        en_val  = _make_rich_text(rr.en,  all_hi_en)  if all_hi_en  else rr.en
        ko_val  = _make_rich_text(rr.ko,  all_hi_ko)  if all_hi_ko  else rr.ko
        en2_val = _make_rich_text(rr.en2, all_hi_en2) if all_hi_en2 else rr.en2
        ko2_val = _make_rich_text(rr.ko2, all_hi_ko2) if all_hi_ko2 else rr.ko2

        for ii, iss in enumerate(rr.issues):
            flat.append(dict(
                rr=rr, iss=iss, is_first=(ii == 0),
                en_val=en_val,   ko_val=ko_val,
                en2_val=en2_val, ko2_val=ko2_val,
                en_err=en_err,   ko_err=ko_err,
                en2_err=en2_err, ko2_err=ko2_err,
            ))

    # ── 데이터 행 기재 ─────────────────────────────────────────
    for seq, row in enumerate(flat, 1):
        ri  = seq + 1          # 헤더가 row 1
        rr  = row["rr"]
        iss = row["iss"]
        fst = row["is_first"]

        def _wcell(col_name, val, *, is_rich=False, font=norm_font, fill=None):
            """
            셀에 값을 기록하고 스타일을 적용한다.
            is_rich=True 이면 CellRichText가 포함될 수 있으므로
            c.font 를 설정하지 않는다 — InlineFont 의 sz 가 보존된다.
            """
            c = ws.cell(ri, COLS[col_name], val)
            c.alignment = TOP
            if fill:
                c.fill = fill
            if not is_rich:
                c.font = font
            return c

        # Index
        _wcell("Index", seq)

        # Report
        _wcell("Report", rr.title)

        # ── 기본 한글명 ──────────────────────────────────────────
        ko_v    = row["ko_val"] if fst else rr.ko
        ko_fill = yell_fill if (fst and row["ko_err"]) else None
        _wcell("기본 한글명",
               ko_v,
               is_rich=isinstance(ko_v, CellRichText),
               fill=ko_fill)

        # ── 기본 영문명 ──────────────────────────────────────────
        en_v    = row["en_val"] if fst else rr.en
        en_fill = yell_fill if (fst and row["en_err"]) else None
        _wcell("기본 영문명",
               en_v,
               is_rich=isinstance(en_v, CellRichText),
               fill=en_fill)

        # ── 표현속성 ─────────────────────────────────────────────
        _wcell("표현속성", rr.lt)

        # ── 표현 한글명 ──────────────────────────────────────────
        ko2_v    = row["ko2_val"] if fst else rr.ko2
        ko2_fill = yell_fill if (fst and row["ko2_err"]) else None
        _wcell("표현 한글명",
               ko2_v,
               is_rich=isinstance(ko2_v, CellRichText),
               fill=ko2_fill)

        # ── 표현 영문명 ──────────────────────────────────────────
        en2_v    = row["en2_val"] if fst else rr.en2
        en2_fill = yell_fill if (fst and row["en2_err"]) else None
        _wcell("표현 영문명",
               en2_v,
               is_rich=isinstance(en2_v, CellRichText),
               fill=en2_fill)

        # ── ErrorType — [AI] 접두사 제거 후 표시 ─────────────────
        etype = re.sub(r"^\[AI\]\s*", "", iss.error_type)
        _wcell("ErrorType", etype, font=err_font)

        # ── ErrorDescription — 정정공시 대상 포맷 적용 ───────────
        desc_val = _format_description(iss.description)
        c = ws.cell(ri, COLS["ErrorDescription"], desc_val)
        c.alignment = TOP
        if not isinstance(desc_val, CellRichText):
            c.font = desc_font

        ws.row_dimensions[ri].height = 42

    # ── 테두리 ─────────────────────────────────────────────────
    last_row = len(flat) + 1
    _apply_full_border(ws, last_row, N)

    # ── ListObject (Excel Table) ────────────────────────────────
    tbl_ref   = f"A1:{get_column_letter(N)}{last_row}"
    tbl       = Table(displayName="InspectionResults", ref=tbl_ref)
    tbl_style = TableStyleInfo(
        name             = "TableStyleLight9",
        showFirstColumn  = False,
        showLastColumn   = False,
        showRowStripes   = False,   # 수동 fill과 충돌 방지
        showColumnStripes= False,
    )
    tbl.tableStyleInfo = tbl_style
    ws.add_table(tbl)

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ─────────────────────────────────────────────────────────────
# File ingestion
# ─────────────────────────────────────────────────────────────
SHEET_CANDIDATES = [
    "XBRLMPMaster", "연결", "별도", "Sheet1", "Data", "재무제표", "MPMaster"
]


def _read_dataframe(path: Path, pwc_encoded: bool = False) -> pd.DataFrame:
    """Read the source xlsx/xlsm and return a normalized DataFrame."""
    # Detect sheet
    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = None
    for candidate in SHEET_CANDIDATES:
        if candidate in xl.sheet_names:
            sheet = candidate
            break
    if sheet is None:
        sheet = xl.sheet_names[0]

    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    # Decode PwC base64 encoded labels if requested
    if pwc_encoded:
        for col in ["ko_label", "en_label", "ko", "en"]:
            if col in df.columns:
                df[col] = df[col].apply(decode_pwc)
    else:
        for col in ["ko_label", "en_label", "ko", "en"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str).str.strip()

    return df


def _has_extra_columns(df: pd.DataFrame) -> bool:
    return all(c in df.columns for c in ["labelTitle", "ko", "en"])


def run_check_bytes(
    file_bytes:  bytes,
    filename:    str,
    company:     str,
    btype:       str,
    pwc_encoded: bool = False,
) -> tuple[bytes, dict]:
    """
    바이트 데이터로부터 XBRL 검사를 수행하고 Excel 보고서를 생성합니다.
    
    Args:
        file_bytes: Excel 파일의 바이트 데이터
        filename: 파일명
        company: 회사명
        btype: 재무제표 구분 ("별도" 또는 "연결")
        pwc_encoded: PwC 인코딩 적용 여부
    
    Returns:
        tuple: (Excel 바이트 데이터, 통계 정보)
    """
    # 임시 파일로 저장
    tmp_path = Path("temp_check.xlsx")
    try:
        with open(tmp_path, "wb") as f:
            f.write(file_bytes)
        
        # DataFrame 읽기
        df = _read_dataframe(tmp_path, pwc_encoded=pwc_encoded)
        has_extra = _has_extra_columns(df)
        
        # ── 1단계: 규칙 기반 검사 ────────────────────────────────
        result_map = {}
        results = []
        issue_cache = {}
        
        for idx, row in df.iterrows():
            title = str(row.get("labelTitle", "") or "").strip() if has_extra else "기본"
            ko = str(row.get("ko_label", "") or "").strip()
            en = str(row.get("en_label", "") or "").strip()
            ko2 = str(row.get("ko", "") or "").strip() if has_extra else ""
            en2 = str(row.get("en", "") or "").strip() if has_extra else ""
            pfx = str(row.get("prefix", "") or "").strip()
            
            cache_key = (en, ko, en2, ko2, pfx, title)
            if cache_key in issue_cache:
                issues = issue_cache[cache_key]
            else:
                issues = detect_issues(en, ko, en2, ko2, prefix=pfx, label_title=title)
                issue_cache[cache_key] = issues
            
            rr = RowResult(title=title, ko=ko, en=en, lt=title, ko2=ko2, en2=en2, issues=issues)
            if issues:  # 문제 있는 행만 결과에 포함
                results.append(rr)
            result_map[idx] = rr
        
        # ── 2단계: Excel 보고서 생성 ──────────────────────────────
        buf = io.BytesIO()
        generate_report(results, buf, company=company, btype=btype, total_rows=len(df))
        buf.seek(0)
        excel_bytes = buf.read()
        
        stats = {
            "total_rows": len(df),
            "issue_count": len(results),
            "n_missing": sum(1 for rr in results for i in rr.issues if i.error_type == "영문명 미기재"),
            "n_violation": sum(1 for rr in results for i in rr.issues if i.error_type == "XBRL 확장 원칙 위배"),
            "n_typo": sum(1 for rr in results for i in rr.issues if i.error_type == "단순 오탈자"),
        }
        return excel_bytes, stats
    
    finally:
        tmp_path.unlink(missing_ok=True)
    """
    Main entry point: run checks on a single file and write the Excel report.
    Returns the output path.
    """
    df = _read_dataframe(path, pwc_encoded=pwc_encoded)
    has_extra = _has_extra_columns(df)

    # ── 1단계: 규칙 기반 검사 ────────────────────────────────
    # row_idx → RowResult 매핑 (AI 병합용)
    result_map: dict[int, RowResult] = {}
    results: list[RowResult] = []

    for idx, row in df.iterrows():
        en    = str(row.get("en_label", "") or "").strip()
        ko    = str(row.get("ko_label", "") or "").strip()
        en2   = str(row.get("en",       "") or "").strip() if has_extra else ""
        ko2   = str(row.get("ko",       "") or "").strip() if has_extra else ""
        lt    = str(row.get("labelTitle","") or "").strip() if has_extra else ""
        title = str(row.get("Title",    "") or "").strip()
        pfx   = str(row.get("prefix",   "") or "").strip()

        issues = detect_issues(en, ko, en2, ko2, prefix=pfx, label_title=lt)
        rr = RowResult(title=title, ko=ko, en=en, lt=lt, ko2=ko2, en2=en2, issues=issues)
        if issues:
            results.append(rr)
        result_map[idx] = rr   # AI 이슈 병합을 위해 항상 저장

    # ── 2단계: Claude API 2차 검토 ───────────────────────────
    n_ai = 0
    if ai_review:
        if not _AI_AVAILABLE:
            print("  [AI] xbrl_ai_reviewer.py 를 찾을 수 없습니다. AI 검토를 건너뜁니다.")
        else:
            entity_rows = prepare_entity_rows(df)
            ai_issues   = review_labels(entity_rows, engine=ai_engine, api_key=api_key, verbose=True)

            for ai_iss in ai_issues:
                rr = result_map.get(ai_iss.row_idx)
                if rr is None:
                    continue

                # Issue 객체로 변환
                if ai_iss.field == "en_label":
                    iss = Issue(
                        error_type   = f"[AI] {ai_iss.error_type}",
                        description  = ai_iss.description,
                        highlight_en = ai_iss.highlight,
                    )
                else:
                    iss = Issue(
                        error_type    = f"[AI] {ai_iss.error_type}",
                        description   = ai_iss.description,
                        highlight_en2 = ai_iss.highlight,
                    )

                rr.issues.append(iss)
                # 아직 결과 목록에 없으면 추가
                if rr not in results:
                    results.append(rr)
                n_ai += 1

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{company}_{btype}_오탈자검사.xlsx"
    generate_report(results, out_path, company, btype, total_rows=len(df))

    n_violations = sum(
        1 for rr in results
        for iss in rr.issues
        if iss.error_type == "XBRL 확장 원칙 위배"
    )
    n_missing = sum(
        1 for rr in results
        for iss in rr.issues
        if iss.error_type == "영문명 미기재"
    )
    n_typo = sum(
        1 for rr in results
        for iss in rr.issues
        if iss.error_type == "단순 오탈자"
    )
    print(
        f"[{company} {btype}] {len(df):,}행 검사 → "
        f"오류항목 {len(results)}건 "
        f"(영문명미기재:{n_missing} / 확장원칙위배:{n_violations} / 오탈자:{n_typo})"
    )
    print(f"  → {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────
def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="XBRL 확장 계정 영문명(en_label) 적절성 검토 도구",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python xbrl_en_label_checker.py LG화학.xlsx --company LG화학 --type 별도
  python xbrl_en_label_checker.py 사조씨푸드.xlsm --pwc-encoded --company 사조씨푸드 --type 연결
  python xbrl_en_label_checker.py *.xlsx --batch --output-dir ./결과
        """,
    )
    p.add_argument("inputs",        nargs="+", help="검사할 xlsx/xlsm 파일 경로 (glob 가능)")
    p.add_argument("--company",     default="",      help="회사명 (단일 파일 처리시)")
    p.add_argument("--type",        default="별도",  help="구분: 별도 | 연결 (단일 파일 처리시)")
    p.add_argument("--output-dir",  default="./결과", help="결과 파일 저장 경로 (기본: ./결과)")
    p.add_argument("--pwc-encoded", action="store_true",
                   help="PwC xlsm 파일의 base64 인코딩된 레이블 디코딩 활성화")
    p.add_argument("--batch",       action="store_true",
                   help="다수 파일 일괄 처리 — 파일명에서 회사명/구분 자동 추출")
    p.add_argument("--ai-review",   action="store_true",
                   help="Claude API 2차 정밀 검토 활성화")
    p.add_argument("--engine",      default="claude",
                   choices=["spellcheck", "claude", "both"],
                   help="2차 검토 엔진 (기본: claude, --ai-review 시 유효)")
    p.add_argument("--api-key",     default="",
                   help="Anthropic API 키 (미지정 시 ANTHROPIC_API_KEY 환경변수)")
    return p


def _infer_meta_from_filename(fname: str) -> tuple[str, str]:
    """
    파일명에서 회사명과 구분(별도/연결)을 추론합니다.
    예: LG화학_2512_별도.xlsx → ("LG화학", "별도")
        XBRL-Wizard_한화솔루션_2512_연결.xlsx → ("한화솔루션", "연결")
    """
    stem = Path(fname).stem
    # 구분 추출
    btype = "별도"
    if "연결" in stem:
        btype = "연결"
    elif "별도" in stem:
        btype = "별도"

    # XBRL-Wizard_ 또는 숫자_ 접두사 제거 후 첫 번째 한글 토큰을 회사명으로
    stem_clean = re.sub(r"^(XBRL-Wizard_|XBRL_Wizard_|\d{8}_)", "", stem)
    parts = re.split(r"[_\-]", stem_clean)
    # 첫 번째 의미 있는 파트 (숫자 아닌 것)
    company = ""
    for part in parts:
        if part and not re.match(r"^\d+$", part) and part not in ("별도", "연결", "v1", "v2", "v3"):
            company = part
            break
    return company or stem, btype


def main() -> None:
    parser = _build_parser()
    args   = parser.parse_args()

    output_dir = Path(args.output_dir)
    all_files: list[Path] = []
    for pattern in args.inputs:
        p = Path(pattern)
        if p.exists():
            all_files.append(p)
        else:
            parent = p.parent if p.parent != Path(".") else Path(".")
            matched = list(parent.glob(p.name))
            all_files.extend(matched if matched else [p])

    if not all_files:
        print("오류: 입력 파일을 찾을 수 없습니다.", file=sys.stderr)
        sys.exit(1)

    for fp in all_files:
        if not fp.exists():
            print(f"경고: 파일을 찾을 수 없음 — {fp}", file=sys.stderr)
            continue

        if args.batch or len(all_files) > 1:
            company, btype = _infer_meta_from_filename(fp.name)
        else:
            company = args.company or fp.stem
            btype   = args.type

        pwc = args.pwc_encoded or fp.suffix.lower() == ".xlsm"

        run_check(
            path        = fp,
            company     = company,
            btype       = btype,
            output_dir  = output_dir,
            pwc_encoded = pwc,
            ai_review   = args.ai_review,
            ai_engine   = args.engine,
            api_key     = args.api_key,
        )


if __name__ == "__main__":
    main()


def run_master_check_bytes(
    file_bytes: bytes,
    filename: str,
) -> tuple:
    """
    마스터 파일용: 모든 행을 검사하여 결과 Excel 바이트를 반환한다.
    마스터 파일은 prefix 컬럼이 없고 모든 행을 검사한다.
    Returns (excel_bytes, stats_dict)
    """
    import io
    import tempfile

    suffix = Path(filename).suffix
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(file_bytes)
        tmp_path = Path(tmp.name)

    try:
        # 마스터 파일은 PwC 인코딩 적용
        df = _read_dataframe(tmp_path, pwc_encoded=True)
        
        results: list[RowResult] = []
        result_map: dict[int, RowResult] = {}
        
        # 캐시를 사용하여 중복 검사 방지
        issue_cache: dict[str, list[Issue]] = {}
        
        for idx, row in df.iterrows():
            en = str(row.get("en_label", "") or "").strip()
            ko = str(row.get("ko_label", "") or "").strip()
            
            # 마스터 파일은 표현 필드 없음
            en2 = ""
            ko2 = ""
            lt = ""
            title = str(row.get("Title", "") or "").strip()
            pfx = ""  # 마스터 파일은 prefix 없음
            
            # 캐시 키 생성 (en_label + ko_label 조합)
            cache_key = f"{en}|{ko}"
            
            if cache_key in issue_cache:
                issues = issue_cache[cache_key]
            else:
                issues = detect_issues(en, ko, en2, ko2, prefix=pfx, label_title=lt)
                issue_cache[cache_key] = issues
            
            rr = RowResult(title=title, ko=ko, en=en, lt=lt, ko2=ko2, en2=en2, issues=issues)
            if issues:  # 마스터 파일은 모든 행을 결과에 포함 (문제 있는 행만 필터링하지 않음)
                results.append(rr)
            result_map[idx] = rr

        buf = io.BytesIO()
        generate_master_report(results, buf, total_rows=len(df))
        buf.seek(0)
        excel_bytes = buf.read()

        stats = {
            "total_rows": len(df),
            "issue_count": len(results),
            "n_missing": sum(1 for rr in results for i in rr.issues if i.error_type == "영문명 미기재"),
            "n_violation": sum(1 for rr in results for i in rr.issues if i.error_type == "XBRL 확장 원칙 위배"),
            "n_typo": sum(1 for rr in results for i in rr.issues if i.error_type == "단순 오탈자"),
        }
        return excel_bytes, stats

    finally:
        tmp_path.unlink(missing_ok=True)


def run_ai_review(excel_bytes: bytes, api_key: str, company: str, btype: str) -> tuple[bytes, dict]:
    """
    AI 2차 검토를 수행합니다.
    
    Args:
        excel_bytes: 검토할 Excel 파일의 바이트 데이터
        api_key: Claude API 키
        company: 회사명
        btype: 재무제표 구분 ("별도" 또는 "연결")
    
    Returns:
        tuple: (AI 검토 결과 Excel 바이트, 통계 정보)
    """
    if not _AI_AVAILABLE:
        raise ImportError("AI 리뷰어 모듈이 설치되지 않았습니다.")
    
    # 임시 파일로 저장
    tmp_path = Path("temp_ai_review.xlsx")
    try:
        with open(tmp_path, "wb") as f:
            f.write(excel_bytes)
        
        # Excel 파일에서 데이터 읽기
        xl = pd.ExcelFile(tmp_path, engine="openpyxl")
        sheet_candidates = ["XBRLMPMaster", "연결", "별도", "Sheet1", "Data", "재무제표"]
        sheet_name = next((s for s in sheet_candidates if s in xl.sheet_names), xl.sheet_names[0])
        df = pd.read_excel(tmp_path, sheet_name=sheet_name, engine="openpyxl")
        
        # PwC 디코딩 적용 (필요시)
        for col in ["ko_label", "en_label", "ko", "en"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str).str.strip()
        
        # Entity rows 준비 및 AI 검토
        entity_rows = prepare_entity_rows(df)
        ai_issues = review_labels(entity_rows, engine="claude", api_key=api_key, verbose=False)
        
        # 기존 워크북 로드
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        
        # AI 검토 시트 추가 또는 업데이트
        if "AI_Review" in wb.sheetnames:
            ws = wb["AI_Review"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("AI_Review")
        
        # 헤더 작성
        headers = ["행 번호", "한글명", "영문명", "AI 검토 결과", "신뢰도", "설명"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name=FONT_NAME, size=FONT_SIZE)
            cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        
        # AI 이슈 작성
        for row, issue in enumerate(ai_issues, 2):
            # 행 번호 찾기 (entity_rows에서 _row_idx 사용)
            row_number = getattr(issue, 'row_idx', issue.row_idx if hasattr(issue, 'row_idx') else row-1)
            
            # 한글명과 영문명 찾기 (entity_rows에서 해당 행 데이터 사용)
            korean_label = ""
            english_label = ""
            for entity_row in entity_rows:
                if entity_row.get("_row_idx") == row_number:
                    korean_label = entity_row.get("ko_label", "")
                    english_label = entity_row.get("en_label", "")
                    break
            
            ws.cell(row=row, column=1, value=row_number + 1)  # 1-based 행 번호
            ws.cell(row=row, column=2, value=korean_label)
            ws.cell(row=row, column=3, value=english_label)
            ws.cell(row=row, column=4, value=issue.error_type)
            ws.cell(row=row, column=5, value="N/A")  # 신뢰도 정보 없음
            ws.cell(row=row, column=6, value=issue.description)
        
        # 열 너비 자동 조정
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 결과 저장
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result_bytes = buf.read()
        
        stats = {
            "ai_issues_count": len(ai_issues),
            "high_confidence": 0,  # 신뢰도 정보 없음
            "medium_confidence": 0,
            "low_confidence": 0,
        }
        
        return result_bytes, stats
    
    finally:
        tmp_path.unlink(missing_ok=True)
